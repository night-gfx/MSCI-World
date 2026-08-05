app_code = r'''
from __future__ import annotations

import json
import os
import tempfile
import threading
import time
from datetime import date, datetime
from pathlib import Path
from typing import Any
from zipfile import BadZipFile

import numpy as np
import pandas as pd
import plotly.graph_objects as go
import yfinance as yf
from dash import Dash, Input, Output, State, ctx, dash_table, dcc, html, no_update
from openpyxl import Workbook, load_workbook
from plotly.subplots import make_subplots


APP_DIR = Path(__file__).resolve().parent
EXCEL_PATH = APP_DIR / "Data.xlsx"

LEGACY_PRICE_SHEET = "FR0014010HV4"
INTRADAY_PRICE_SHEET = f"{LEGACY_PRICE_SHEET} (5 min)"
DAILY_PRICE_SHEET = f"{LEGACY_PRICE_SHEET} (Daily)"
SETUP_SHEET = "Setup"

INSTRUMENT_NAME = "Amundi MSCI World (2x) Leveraged UCITS ETF Acc"
INSTRUMENT_ISIN = "FR0014010HV4"
YAHOO_TICKER = "LWLD.PA"
DEFAULT_INSTRUMENT_KEY = "amundi_2x"
INSTRUMENTS = {
    DEFAULT_INSTRUMENT_KEY: {
        "name": INSTRUMENT_NAME,
        "isin": INSTRUMENT_ISIN,
        "ticker": YAHOO_TICKER,
    },
    "ishares_msci_world": {
        "name": "iShares MSCI World UCITS ETF USD (Dist)",
        "isin": "IE00B0M62Q58",
        "ticker": "IQQW.DE",
    },
    "co2_allowances": {
        "name": "SparkChange Physical Carbon EUA ETC",
        "isin": "XS2353177293",
        "ticker": "CO2.L",
    },
}
INTRADAY_INTERVAL = "5m"
INTRADAY_PERIOD = "60d"
DAILY_INTERVAL = "1d"
DAILY_PERIOD = "max"
MARKET_REFRESH_SECONDS = 300
PRICE_COLUMNS = [
    "Timestamp",
    "Open",
    "High",
    "Low",
    "Close",
    "Adj Close",
    "Volume",
    "Ticker",
    "ISIN",
]

DEFAULT_SETTINGS = {
    "regression_6m_days": 182,
    "regression_1y_days": 365,
    "regression_2y_days": 730,
    "regression_5y_days": 1825,
    "show_regression_6m": True,
    "show_regression_1y": True,
    "show_regression_2y": True,
    "show_regression_5y": True,
    "show_regression_max": False,
    "bollinger_window": 20,
    "bollinger_std": 2.0,
    "show_bollinger": True,
    "kalman_process_variance": 1.0,
    "kalman_measurement_variance": 25.0,
    "show_kalman": True,
}

OBSOLETE_SETTINGS = {
    "regression_custom_days",
    "show_regression_custom",
}

SETTINGS_START_ROW = 1
LEGACY_TRADES_HEADER_ROW = 15
TRADES_HEADER_ROW = 20

TRADE_COLUMNS = [
    "Trade_ID",
    "Entry_Date",
    "Exit_Date",
    "Direction",
    "Quantity",
    "Entry_Price",
    "Exit_Price",
    "Fees",
    "Notes",
    "Status",
    "Holding_Days",
    "Gross_Return_Pct",
    "Net_PnL",
    "Net_Return_Pct",
]

MARKET_DATA_CACHE: dict[str, Any] = {
    "signature": None,
    "data": None,
}
DAILY_MARKET_DATA_CACHE: dict[str, Any] = {
    "signature": None,
    "data": None,
}
INSTRUMENT_MARKET_CACHES: dict[tuple[str, str], dict[str, Any]] = {}
LAST_PRICE_CACHE: dict[str, Any] = {
    "fetched_monotonic": 0.0,
    "timestamp": None,
    "price": None,
}
INSTRUMENT_LAST_PRICE_CACHES: dict[str, dict[str, Any]] = {
    DEFAULT_INSTRUMENT_KEY: LAST_PRICE_CACHE,
}
TRADES_CACHE: dict[str, Any] = {
    "signature": None,
    "data": None,
}
REGRESSION_CACHE: dict[str, Any] = {
    "signature": None,
    "series": {},
}
KALMAN_CACHE: dict[str, Any] = {
    "signature": None,
    "series": {},
}
MARKET_REFRESH_LOCK = threading.Lock()
LAST_MARKET_REFRESH_MONOTONIC = {
    INTRADAY_PRICE_SHEET: 0.0,
    DAILY_PRICE_SHEET: 0.0,
}
BOLLINGER_CACHE: dict[str, Any] = {
    "signature": None,
    "series": {},
}
DEFAULT_KALMAN_PROCESS_VARIANCE = 1.0
DEFAULT_KALMAN_MEASUREMENT_VARIANCE = 25.0


def instrument_config(instrument_key: str | None = None) -> dict[str, str]:
    return INSTRUMENTS.get(instrument_key or DEFAULT_INSTRUMENT_KEY, INSTRUMENTS[DEFAULT_INSTRUMENT_KEY])


def instrument_sheet_name(instrument_key: str | None, frequency: str) -> str:
    config = instrument_config(instrument_key)
    suffix = "5 min" if frequency == "intraday" else "Daily"
    return f"{config['isin']} ({suffix})"


def instrument_market_cache(instrument_key: str | None, frequency: str) -> dict[str, Any]:
    key = instrument_key or DEFAULT_INSTRUMENT_KEY
    if key == DEFAULT_INSTRUMENT_KEY:
        return MARKET_DATA_CACHE if frequency == "intraday" else DAILY_MARKET_DATA_CACHE
    return INSTRUMENT_MARKET_CACHES.setdefault(
        (key, frequency), {"signature": None, "data": None}
    )


def instrument_last_price_cache(instrument_key: str | None) -> dict[str, Any]:
    key = instrument_key or DEFAULT_INSTRUMENT_KEY
    return INSTRUMENT_LAST_PRICE_CACHES.setdefault(
        key, {"fetched_monotonic": 0.0, "timestamp": None, "price": None}
    )


def ensure_workbook_exists() -> None:
    if not EXCEL_PATH.exists():
        raise FileNotFoundError(
            f"Die Excel-Datei {EXCEL_PATH.name} fehlt. "
            "Sie wird nicht automatisch erzeugt, damit keine Kursdaten ersetzt werden."
        )

    try:
        wb = load_workbook(EXCEL_PATH, read_only=True, data_only=False)
        wb.close()
    except (BadZipFile, OSError, ValueError) as exc:
        raise RuntimeError(
            f"Die Excel-Datei {EXCEL_PATH.name} ist beschädigt oder ungültig. "
            "Die Datei bleibt unverändert; bitte stelle eine gültige Version wieder her."
        ) from exc


def _style_price_history_sheet(ws) -> None:
    from openpyxl.styles import Font, PatternFill

    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)

    widths = {
        "A": 21,
        "B": 13,
        "C": 13,
        "D": 13,
        "E": 13,
        "F": 13,
        "G": 14,
        "H": 12,
        "I": 17,
    }
    for column, width in widths.items():
        ws.column_dimensions[column].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:I{max(ws.max_row, 1)}"


def ensure_price_history_structures(instrument_key: str | None = None) -> None:
    ensure_workbook_exists()
    selected_key = instrument_key or DEFAULT_INSTRUMENT_KEY
    intraday_sheet = instrument_sheet_name(selected_key, "intraday")
    daily_sheet = instrument_sheet_name(selected_key, "daily")
    wb = load_workbook(EXCEL_PATH)
    changed = False
    try:
        if (
            selected_key == DEFAULT_INSTRUMENT_KEY
            and intraday_sheet not in wb.sheetnames
            and LEGACY_PRICE_SHEET in wb.sheetnames
        ):
            wb[LEGACY_PRICE_SHEET].title = intraday_sheet
            changed = True

        for sheet_name in (intraday_sheet, daily_sheet):
            if sheet_name not in wb.sheetnames:
                ws = wb.create_sheet(sheet_name)
                ws.append(PRICE_COLUMNS)
                _style_price_history_sheet(ws)
                changed = True
                continue

            ws = wb[sheet_name]
            headers = [ws.cell(row=1, column=index).value for index in range(1, 10)]
            if all(value is None for value in headers):
                ws.append(PRICE_COLUMNS)
                _style_price_history_sheet(ws)
                changed = True
            elif headers != PRICE_COLUMNS:
                raise RuntimeError(
                    f"Das Sheet '{sheet_name}' hat unerwartete Spalten. "
                    f"Erwartet: {', '.join(PRICE_COLUMNS)}."
                )

        desired_tail = [intraday_sheet, daily_sheet]
        if wb.sheetnames[-2:] != desired_tail:
            for sheet_name in desired_tail:
                ws = wb[sheet_name]
                wb.move_sheet(ws, offset=len(wb.worksheets) - 1 - wb.index(ws))
            changed = True

        if changed:
            atomic_save_workbook(wb)
        else:
            wb.close()
    except Exception:
        wb.close()
        raise


def _download_prices(
    ticker: str, isin: str, period: str, interval: str
) -> pd.DataFrame:
    downloaded = yf.download(
        ticker,
        period=period,
        interval=interval,
        auto_adjust=False,
        prepost=False,
        progress=False,
        threads=False,
    )
    if downloaded.empty:
        raise RuntimeError(f"Yahoo Finance lieferte keine Daten für {ticker}.")

    if isinstance(downloaded.columns, pd.MultiIndex):
        downloaded.columns = downloaded.columns.get_level_values(0)
    downloaded = downloaded.reset_index()
    timestamp_column = "Datetime" if "Datetime" in downloaded.columns else downloaded.columns[0]
    downloaded = downloaded.rename(columns={timestamp_column: "Timestamp"})
    downloaded["Timestamp"] = pd.to_datetime(downloaded["Timestamp"], errors="coerce")
    if downloaded["Timestamp"].dt.tz is not None:
        downloaded["Timestamp"] = (
            downloaded["Timestamp"].dt.tz_convert("Europe/Berlin").dt.tz_localize(None)
        )

    if "Adj Close" not in downloaded.columns:
        downloaded["Adj Close"] = downloaded.get("Close")
    downloaded["Ticker"] = ticker
    downloaded["ISIN"] = isin
    for column in ["Open", "High", "Low", "Close", "Adj Close", "Volume"]:
        downloaded[column] = pd.to_numeric(downloaded.get(column), errors="coerce")
    downloaded = downloaded[PRICE_COLUMNS].dropna(subset=["Timestamp", "Adj Close"])
    return downloaded.sort_values("Timestamp").drop_duplicates("Timestamp", keep="last")


def refresh_market_history(
    instrument_key: str,
    sheet_name: str,
    period: str,
    interval: str,
    cache: dict[str, Any],
    force: bool = False,
) -> int:
    with MARKET_REFRESH_LOCK:
        last_refresh = LAST_MARKET_REFRESH_MONOTONIC.get(sheet_name, 0.0)
        elapsed = time.monotonic() - last_refresh
        if not force and last_refresh and elapsed < MARKET_REFRESH_SECONDS:
            return 0

        config = instrument_config(instrument_key)
        fresh = _download_prices(
            ticker=config["ticker"],
            isin=config["isin"],
            period=period,
            interval=interval,
        )
        LAST_MARKET_REFRESH_MONOTONIC[sheet_name] = time.monotonic()
        ensure_price_history_structures(instrument_key)

        wb = load_workbook(EXCEL_PATH)
        try:
            ws = wb[sheet_name]
            existing_timestamps = {
                pd.Timestamp(value).to_pydatetime()
                for (value,) in ws.iter_rows(min_row=2, max_col=1, values_only=True)
                if value is not None
            }
            new_rows = []
            for row in fresh.itertuples(index=False, name=None):
                timestamp = pd.Timestamp(row[0]).to_pydatetime()
                if timestamp in existing_timestamps:
                    continue
                new_rows.append([timestamp, *row[1:]])

            for row in new_rows:
                ws.append(row)
            if new_rows:
                for cell in ws["A"][1:]:
                    cell.number_format = "yyyy-mm-dd hh:mm"
                for column in ("B", "C", "D", "E", "F"):
                    for cell in ws[column][1:]:
                        cell.number_format = "0.0000"
                for cell in ws["G"][1:]:
                    cell.number_format = "#,##0"
                ws.auto_filter.ref = f"A1:I{ws.max_row}"
                atomic_save_workbook(wb)
            else:
                wb.close()
        except Exception:
            wb.close()
            raise

        if new_rows:
            cache["signature"] = None
            cache["data"] = None
        return len(new_rows)


def _write_default_setup(ws_setup) -> None:
    ws_setup["A1"] = "Setting"
    ws_setup["B1"] = "Value"

    row = 2
    for key, value in DEFAULT_SETTINGS.items():
        ws_setup.cell(row=row, column=1, value=key)
        ws_setup.cell(row=row, column=2, value=value)
        row += 1

    for col_idx, name in enumerate(TRADE_COLUMNS, start=1):
        ws_setup.cell(row=TRADES_HEADER_ROW, column=col_idx, value=name)


def ensure_setup_structure() -> None:
    ensure_workbook_exists()
    try:
        wb = load_workbook(EXCEL_PATH)
    except (BadZipFile, OSError, ValueError) as exc:
        raise RuntimeError(
            f"Die Excel-Datei {EXCEL_PATH.name} konnte nicht geöffnet werden. "
            "Sie wurde nicht verändert."
        ) from exc

    if SETUP_SHEET not in wb.sheetnames:
        ws_setup = wb.create_sheet(SETUP_SHEET)
        _write_default_setup(ws_setup)
        atomic_save_workbook(wb)
        return

    ws = wb[SETUP_SHEET]
    setup_changed = False

    if (
        ws.cell(row=LEGACY_TRADES_HEADER_ROW, column=1).value == "Trade_ID"
        and ws.cell(row=TRADES_HEADER_ROW, column=1).value != "Trade_ID"
    ):
        ws.move_range(
            f"A{LEGACY_TRADES_HEADER_ROW}:N{ws.max_row}",
            rows=TRADES_HEADER_ROW - LEGACY_TRADES_HEADER_ROW,
            cols=0,
        )
        setup_changed = True

    if ws["A1"].value != "Setting" or ws["B1"].value != "Value":
        setup_is_empty = not any(
            cell.value is not None
            for row in ws.iter_rows()
            for cell in row
        )
        if setup_is_empty:
            _write_default_setup(ws)
            setup_changed = True
        else:
            wb.close()
            raise RuntimeError(
                f"Das vorhandene Sheet '{SETUP_SHEET}' ist kein Konfigurations-Sheet "
                "(erwartet werden 'Setting' und 'Value' in A1/B1). "
                "Es bleibt vollständig unverändert."
            )

    for row in range(2, TRADES_HEADER_ROW):
        key = ws.cell(row=row, column=1).value
        if key in OBSOLETE_SETTINGS:
            ws.cell(row=row, column=1).value = None
            ws.cell(row=row, column=2).value = None
            setup_changed = True

    existing_settings = {}
    for row in range(2, TRADES_HEADER_ROW):
        key = ws.cell(row=row, column=1).value
        if key:
            existing_settings[str(key)] = row

    next_row = 2
    for key, value in DEFAULT_SETTINGS.items():
        if key in existing_settings:
            continue
        while ws.cell(row=next_row, column=1).value is not None:
            next_row += 1
        if next_row >= TRADES_HEADER_ROW:
            raise RuntimeError(
                "Im Setup-Sheet ist kein Platz mehr für weitere Einstellungen "
                f"oberhalb von Zeile {TRADES_HEADER_ROW}."
            )
        ws.cell(row=next_row, column=1, value=key)
        ws.cell(row=next_row, column=2, value=value)
        setup_changed = True

    if ws.cell(row=TRADES_HEADER_ROW, column=1).value != "Trade_ID":
        for col_idx, name in enumerate(TRADE_COLUMNS, start=1):
            ws.cell(row=TRADES_HEADER_ROW, column=col_idx, value=name)
        setup_changed = True

    if setup_changed:
        atomic_save_workbook(wb)
    else:
        wb.close()


def _load_market_data(
    instrument_key: str,
    sheet_name: str,
    cache: dict[str, Any],
    description: str,
) -> pd.DataFrame:
    ensure_price_history_structures(instrument_key)

    if EXCEL_PATH.exists():
        stat = EXCEL_PATH.stat()
        signature = (stat.st_mtime_ns, stat.st_size)
        if cache["signature"] == signature:
            cached = cache["data"]
            if isinstance(cached, pd.DataFrame):
                return cached

    ensure_workbook_exists()

    try:
        df = pd.read_excel(EXCEL_PATH, sheet_name=sheet_name)
    except ValueError as exc:
        raise RuntimeError(
            f"Das Sheet '{sheet_name}' fehlt in {EXCEL_PATH.name}."
        ) from exc

    if df.empty or df.shape[1] < 2:
        raise RuntimeError(
            f"Das Sheet '{sheet_name}' besitzt keine gültigen Kursdaten. "
            "Bitte prüfe die Yahoo-Verbindung und versuche es erneut."
        )

    date_col = "Timestamp" if "Timestamp" in df.columns else df.columns[0]
    # Both histories are evaluated using adjusted closes. Close remains a
    # compatibility fallback for an older workbook.
    if "Adj Close" in df.columns:
        price_col = "Adj Close"
    elif "Close" in df.columns:
        price_col = "Close"
    else:
        price_col = df.columns[1]

    out = df[[date_col, price_col]].copy()
    out.columns = ["Date", "Price"]
    out["Date"] = pd.to_datetime(out["Date"], dayfirst=True, errors="coerce")
    out["Price"] = pd.to_numeric(out["Price"], errors="coerce")
    out = out.dropna(subset=["Date", "Price"])
    out = out.sort_values("Date").drop_duplicates("Date", keep="last").reset_index(drop=True)

    if out.empty:
        raise RuntimeError(
            f"Im Sheet '{sheet_name}' wurden keine gültigen {description} gefunden."
        )

    stat = EXCEL_PATH.stat()
    source_signature = (stat.st_mtime_ns, stat.st_size)
    out.attrs["source_signature"] = source_signature
    cache["signature"] = source_signature
    cache["data"] = out
    return out


def load_market_data(instrument_key: str = DEFAULT_INSTRUMENT_KEY) -> pd.DataFrame:
    """Load adjusted 5-minute prices for Trading Analytics and trade capture."""
    return _load_market_data(
        instrument_key=instrument_key,
        sheet_name=instrument_sheet_name(instrument_key, "intraday"),
        cache=instrument_market_cache(instrument_key, "intraday"),
        description="5-Minuten-Daten",
    )


def load_daily_market_data(instrument_key: str = DEFAULT_INSTRUMENT_KEY) -> pd.DataFrame:
    """Load the maximum adjusted daily history for Trading Tools."""
    return _load_market_data(
        instrument_key=instrument_key,
        sheet_name=instrument_sheet_name(instrument_key, "daily"),
        cache=instrument_market_cache(instrument_key, "daily"),
        description="Tagesdaten",
    )


def load_last_price(
    instrument_key: str = DEFAULT_INSTRUMENT_KEY, force: bool = False
) -> tuple[pd.Timestamp, float]:
    """Return Yahoo's latest quote without changing the 5-minute history."""
    cache = instrument_last_price_cache(instrument_key)
    elapsed = time.monotonic() - float(cache["fetched_monotonic"] or 0.0)
    cached_timestamp = cache["timestamp"]
    cached_price = cache["price"]
    if (
        cached_timestamp is not None
        and cached_price is not None
        and elapsed < MARKET_REFRESH_SECONDS
        and not force
    ):
        return pd.Timestamp(cached_timestamp), float(cached_price)

    config = instrument_config(instrument_key)
    quote = yf.Ticker(config["ticker"])
    price = pd.to_numeric(quote.fast_info.get("last_price"), errors="coerce")
    timestamp = pd.Timestamp.now(tz="Europe/Berlin").tz_localize(None)
    if pd.isna(price) or not np.isfinite(float(price)) or float(price) <= 0:
        fallback = _download_prices(
            ticker=config["ticker"],
            isin=config["isin"],
            period="1d",
            interval="1m",
        )
        if fallback.empty:
            fallback = _download_prices(
                ticker=config["ticker"],
                isin=config["isin"],
                period="5d",
                interval="1d",
            )
        if fallback.empty:
            raise RuntimeError(f"Yahoo Finance lieferte keinen Last Price für {config['ticker']}.")
        latest = fallback.iloc[-1]
        timestamp = pd.Timestamp(latest["Timestamp"])
        price = float(latest["Adj Close"])
    cache.update(
        {
            "fetched_monotonic": time.monotonic(),
            "timestamp": timestamp,
            "price": float(price),
        }
    )
    return timestamp, float(price)


def load_trading_tools_data(
    instrument_key: str = DEFAULT_INSTRUMENT_KEY,
) -> pd.DataFrame:
    """Add the current quote only to the data used by Trading Tools."""
    history = load_daily_market_data(instrument_key)
    cache = instrument_last_price_cache(instrument_key)
    quote_timestamp = cache["timestamp"]
    last_price = cache["price"]
    if quote_timestamp is None or last_price is None:
        return history

    tools_data = history.copy()
    tools_data["Date"] = pd.to_datetime(tools_data["Date"]).dt.normalize()
    latest_history_timestamp = pd.Timestamp(tools_data["Date"].max()).normalize()
    quote_timestamp = max(
        pd.Timestamp(quote_timestamp).normalize(),
        latest_history_timestamp,
    )
    live_row = pd.DataFrame({"Date": [quote_timestamp], "Price": [last_price]})
    tools_data = pd.concat([tools_data, live_row], ignore_index=True)
    tools_data = (
        tools_data.sort_values("Date")
        .drop_duplicates("Date", keep="last")
        .reset_index(drop=True)
    )
    tools_data.attrs["source_signature"] = (
        history.attrs.get("source_signature"),
        quote_timestamp.value,
        last_price,
    )
    return tools_data


def _coerce_bool(value: Any, default: bool) -> bool:
    if isinstance(value, bool):
        return value
    if value is None:
        return default
    if isinstance(value, (int, float)):
        return bool(value)
    return str(value).strip().lower() in {"1", "true", "ja", "yes", "y"}


def _positive_float_input(value: Any, default: float) -> float:
    try:
        parsed = float(str(value).strip().replace(",", "."))
    except (TypeError, ValueError):
        return float(default)
    return parsed if np.isfinite(parsed) and parsed > 0 else float(default)


def _positive_int_input(value: Any, default: int, minimum: int = 1) -> int:
    parsed = int(round(_positive_float_input(value, float(default))))
    return max(parsed, minimum)


def load_settings() -> dict[str, Any]:
    ensure_setup_structure()
    wb = load_workbook(EXCEL_PATH, data_only=True)
    ws = wb[SETUP_SHEET]

    settings = DEFAULT_SETTINGS.copy()
    for row in range(2, TRADES_HEADER_ROW):
        key = ws.cell(row=row, column=1).value
        value = ws.cell(row=row, column=2).value
        if key in settings and value is not None:
            settings[key] = value

    settings["regression_6m_days"] = int(settings["regression_6m_days"])
    settings["regression_1y_days"] = int(settings["regression_1y_days"])
    settings["regression_2y_days"] = int(settings["regression_2y_days"])
    settings["regression_5y_days"] = int(settings["regression_5y_days"])
    settings["bollinger_window"] = int(settings["bollinger_window"])
    settings["bollinger_std"] = float(settings["bollinger_std"])
    settings["kalman_process_variance"] = float(settings["kalman_process_variance"])
    settings["kalman_measurement_variance"] = float(settings["kalman_measurement_variance"])

    for key in (
        "show_regression_6m",
        "show_regression_1y",
        "show_regression_2y",
        "show_regression_5y",
        "show_regression_max",
        "show_bollinger",
        "show_kalman",
    ):
        settings[key] = _coerce_bool(settings[key], DEFAULT_SETTINGS[key])

    return settings


def atomic_save_workbook(wb) -> None:
    temp_fd, temp_name = tempfile.mkstemp(
        prefix="data_", suffix=".xlsx", dir=str(APP_DIR)
    )
    os.close(temp_fd)

    try:
        wb.save(temp_name)
        os.replace(temp_name, EXCEL_PATH)
    except PermissionError as exc:
        if os.path.exists(temp_name):
            os.remove(temp_name)
        raise PermissionError(
            "Die Excel-Datei ist vermutlich in Excel geöffnet. Bitte schließen und erneut speichern."
        ) from exc
    finally:
        if os.path.exists(temp_name):
            os.remove(temp_name)


def save_settings(settings: dict[str, Any]) -> None:
    ensure_setup_structure()
    wb = load_workbook(EXCEL_PATH)
    ws = wb[SETUP_SHEET]

    row_by_key = {}
    for row in range(2, TRADES_HEADER_ROW):
        key = ws.cell(row=row, column=1).value
        if key:
            row_by_key[str(key)] = row

    for key, value in settings.items():
        row = row_by_key.get(key)
        if row is None:
            raise RuntimeError(f"Einstellung '{key}' fehlt im Setup-Sheet.")
        ws.cell(row=row, column=2, value=value)

    atomic_save_workbook(wb)


def load_trades() -> pd.DataFrame:
    if EXCEL_PATH.exists():
        stat = EXCEL_PATH.stat()
        signature = (stat.st_mtime_ns, stat.st_size)
        if TRADES_CACHE["signature"] == signature:
            cached = TRADES_CACHE["data"]
            if isinstance(cached, pd.DataFrame):
                return cached.copy()

    wb = load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    try:
        if SETUP_SHEET not in wb.sheetnames:
            raise RuntimeError(f"Das Sheet '{SETUP_SHEET}' fehlt in {EXCEL_PATH.name}.")
        ws = wb[SETUP_SHEET]
        rows = [
            list(values)
            for values in ws.iter_rows(
                min_row=TRADES_HEADER_ROW + 1,
                max_col=len(TRADE_COLUMNS),
                values_only=True,
            )
            if not all(value is None for value in values)
        ]
    finally:
        wb.close()

    trades = pd.DataFrame(rows, columns=TRADE_COLUMNS)
    if trades.empty:
        stat = EXCEL_PATH.stat()
        TRADES_CACHE["signature"] = (stat.st_mtime_ns, stat.st_size)
        TRADES_CACHE["data"] = trades.copy()
        return trades

    for col in ["Entry_Date", "Exit_Date"]:
        trades[col] = pd.to_datetime(trades[col], errors="coerce")

    for col in [
        "Quantity",
        "Entry_Price",
        "Exit_Price",
        "Fees",
        "Holding_Days",
        "Gross_Return_Pct",
        "Net_PnL",
        "Net_Return_Pct",
    ]:
        trades[col] = pd.to_numeric(trades[col], errors="coerce")

    stat = EXCEL_PATH.stat()
    TRADES_CACHE["signature"] = (stat.st_mtime_ns, stat.st_size)
    TRADES_CACHE["data"] = trades.copy()
    return trades


def nearest_price(df: pd.DataFrame, target_date: pd.Timestamp) -> tuple[pd.Timestamp, float]:
    idx = (df["Date"] - target_date).abs().idxmin()
    row = df.loc[idx]
    return pd.Timestamp(row["Date"]), float(row["Price"])


def calculate_trade_metrics(
    entry_date: pd.Timestamp,
    exit_date: pd.Timestamp | None,
    direction: str,
    quantity: float,
    entry_price: float,
    exit_price: float | None,
    fees: float,
) -> dict[str, Any]:
    status = "Open" if exit_date is None or exit_price is None else "Closed"

    if status == "Open":
        return {
            "Status": status,
            "Holding_Days": None,
            "Gross_Return_Pct": None,
            "Net_PnL": None,
            "Net_Return_Pct": None,
        }

    sign = 1.0 if direction == "Long" else -1.0
    gross_pnl = sign * (exit_price - entry_price) * quantity
    invested_capital = abs(entry_price * quantity)
    net_pnl = gross_pnl - fees
    gross_return = sign * (exit_price / entry_price - 1.0)
    net_return = net_pnl / invested_capital if invested_capital else np.nan

    return {
        "Status": status,
        "Holding_Days": int((exit_date - entry_date).days),
        "Gross_Return_Pct": gross_return * 100.0,
        "Net_PnL": net_pnl,
        "Net_Return_Pct": net_return * 100.0,
    }


def append_trade(
    market_df: pd.DataFrame,
    entry_date_value: str,
    exit_date_value: str | None,
    direction: str,
    quantity: float,
    entry_price_value: float | None,
    exit_price_value: float | None,
    fees: float,
    notes: str,
) -> str:
    entry_date = pd.Timestamp(entry_date_value)
    exit_date = pd.Timestamp(exit_date_value) if exit_date_value else None

    actual_entry_date, market_entry_price = nearest_price(market_df, entry_date)
    entry_price = float(entry_price_value) if entry_price_value is not None else market_entry_price

    actual_exit_date = None
    exit_price = None
    if exit_date is not None:
        actual_exit_date, market_exit_price = nearest_price(market_df, exit_date)
        exit_price = float(exit_price_value) if exit_price_value is not None else market_exit_price

    if actual_exit_date is not None and actual_exit_date < actual_entry_date:
        raise ValueError("Das Exit-Datum darf nicht vor dem Entry-Datum liegen.")

    trade_metrics = calculate_trade_metrics(
        entry_date=actual_entry_date,
        exit_date=actual_exit_date,
        direction=direction,
        quantity=float(quantity),
        entry_price=entry_price,
        exit_price=exit_price,
        fees=float(fees),
    )

    ensure_setup_structure()
    wb = load_workbook(EXCEL_PATH)
    ws = wb[SETUP_SHEET]

    existing_ids = []
    for row_idx in range(TRADES_HEADER_ROW + 1, ws.max_row + 1):
        value = ws.cell(row=row_idx, column=1).value
        if isinstance(value, (int, float)):
            existing_ids.append(int(value))
    trade_id = max(existing_ids, default=0) + 1

    next_row = max(ws.max_row + 1, TRADES_HEADER_ROW + 1)
    row_data = {
        "Trade_ID": trade_id,
        "Entry_Date": actual_entry_date.to_pydatetime(),
        "Exit_Date": actual_exit_date.to_pydatetime() if actual_exit_date is not None else None,
        "Direction": direction,
        "Quantity": float(quantity),
        "Entry_Price": entry_price,
        "Exit_Price": exit_price,
        "Fees": float(fees),
        "Notes": notes,
        **trade_metrics,
    }

    for col_idx, col_name in enumerate(TRADE_COLUMNS, start=1):
        ws.cell(row=next_row, column=col_idx, value=row_data.get(col_name))

    atomic_save_workbook(wb)

    if actual_entry_date.date() != entry_date.date():
        entry_note = f" Entry wurde auf den nächsten verfügbaren Handelstag {actual_entry_date.date()} gelegt."
    else:
        entry_note = ""

    if exit_date is not None and actual_exit_date is not None and actual_exit_date.date() != exit_date.date():
        exit_note = f" Exit wurde auf den nächsten verfügbaren Handelstag {actual_exit_date.date()} gelegt."
    else:
        exit_note = ""

    return f"Trade {trade_id} gespeichert.{entry_note}{exit_note}"


def delete_trade(trade_id: Any) -> str:
    ensure_setup_structure()
    wb = load_workbook(EXCEL_PATH)
    try:
        ws = wb[SETUP_SHEET]
        target = str(trade_id).strip()
        target_row = None
        for row_idx in range(TRADES_HEADER_ROW + 1, ws.max_row + 1):
            stored_id = ws.cell(row=row_idx, column=1).value
            if stored_id is not None and str(stored_id).strip() == target:
                target_row = row_idx
                break
            try:
                if stored_id is not None and float(stored_id) == float(trade_id):
                    target_row = row_idx
                    break
            except (TypeError, ValueError):
                pass
        if target_row is None:
            raise ValueError(f"Trade {trade_id} wurde nicht gefunden.")
        ws.delete_rows(target_row, 1)
        atomic_save_workbook(wb)
    finally:
        wb.close()

    TRADES_CACHE["signature"] = None
    TRADES_CACHE["data"] = None
    return f"Trade {trade_id} wurde gelöscht."


def linear_regression_series(df: pd.DataFrame, window_days: int) -> pd.DataFrame:
    end_date = df["Date"].max()
    start_date = end_date - pd.Timedelta(days=int(window_days))
    window = df[df["Date"] >= start_date].copy()

    if len(window) < 2:
        return pd.DataFrame(columns=["Date", "Regression"])

    # Intraday-safe time axis: keep the fractional part of a day.  Using
    # ``.dt.days`` collapsed every point of one trading day onto the same x value.
    x = (
        (window["Date"] - window["Date"].min()).dt.total_seconds()
        / 86_400.0
    ).to_numpy(dtype=float)
    y = window["Price"].to_numpy(dtype=float)
    slope, intercept = np.polyfit(x, y, deg=1)
    window["Regression"] = intercept + slope * x
    return window[["Date", "Regression"]]


def cached_linear_regression_series(
    df: pd.DataFrame,
    window_days: int,
) -> pd.DataFrame:
    source_signature = df.attrs.get("source_signature")
    if source_signature is None:
        return linear_regression_series(df, window_days)

    if REGRESSION_CACHE["signature"] != source_signature:
        REGRESSION_CACHE["signature"] = source_signature
        REGRESSION_CACHE["series"] = {}

    cache_key = int(window_days)
    cached = REGRESSION_CACHE["series"].get(cache_key)
    if isinstance(cached, pd.DataFrame):
        return cached

    regression = linear_regression_series(df, cache_key)
    REGRESSION_CACHE["series"][cache_key] = regression
    return regression


def cached_bollinger_series(
    df: pd.DataFrame,
    window: int,
    std_mult: float,
) -> pd.DataFrame:
    source_signature = df.attrs.get("source_signature")
    cache_key = (int(window), float(std_mult))
    if source_signature is not None:
        if BOLLINGER_CACHE["signature"] != source_signature:
            BOLLINGER_CACHE["signature"] = source_signature
            BOLLINGER_CACHE["series"] = {}
        cached = BOLLINGER_CACHE["series"].get(cache_key)
        if isinstance(cached, pd.DataFrame):
            return cached

    result = df[["Date", "Price"]].copy()
    result["BB_Middle"] = result["Price"].rolling(window=window).mean()
    rolling_std = result["Price"].rolling(window=window).std(ddof=0)
    result["BB_Upper"] = result["BB_Middle"] + std_mult * rolling_std
    result["BB_Lower"] = result["BB_Middle"] - std_mult * rolling_std
    if source_signature is not None:
        BOLLINGER_CACHE["series"][cache_key] = result
    return result


def cached_kalman_filter_series(
    df: pd.DataFrame,
    process_variance: float,
    measurement_variance: float,
) -> pd.DataFrame:
    process_variance = max(float(process_variance), 1e-9)
    measurement_variance = max(float(measurement_variance), 1e-9)
    source_signature = df.attrs.get("source_signature")
    cache_key = (process_variance, measurement_variance)

    if source_signature is not None:
        if KALMAN_CACHE["signature"] != source_signature:
            KALMAN_CACHE["signature"] = source_signature
            KALMAN_CACHE["series"] = {}
        cached = KALMAN_CACHE["series"].get(cache_key)
        if isinstance(cached, pd.DataFrame):
            return cached

    dates = pd.to_datetime(df["Date"]).to_numpy()
    prices = df["Price"].to_numpy(dtype=float)
    estimates = np.empty(len(prices), dtype=float)
    velocities = np.empty(len(prices), dtype=float)
    state = np.array([float(prices[0]), 0.0], dtype=float)
    covariance = np.diag([measurement_variance, process_variance])
    observation_matrix = np.array([1.0, 0.0], dtype=float)
    identity = np.eye(2, dtype=float)

    estimates[0] = state[0]
    velocities[0] = state[1]
    for index in range(1, len(prices)):
        delta_days = max(
            float((dates[index] - dates[index - 1]) / np.timedelta64(1, "D")),
            1.0 / 1440.0,
        )
        transition = np.array(
            [[1.0, delta_days], [0.0, 1.0]], dtype=float
        )
        process_noise = process_variance * np.array(
            [
                [delta_days**4 / 4.0, delta_days**3 / 2.0],
                [delta_days**3 / 2.0, delta_days**2],
            ],
            dtype=float,
        )

        state = transition @ state
        covariance = transition @ covariance @ transition.T + process_noise
        innovation = float(prices[index] - observation_matrix @ state)
        innovation_variance = float(
            observation_matrix @ covariance @ observation_matrix.T
            + measurement_variance
        )
        kalman_gain = covariance @ observation_matrix / innovation_variance
        state = state + kalman_gain * innovation
        correction = identity - np.outer(kalman_gain, observation_matrix)
        covariance = (
            correction @ covariance @ correction.T
            + np.outer(kalman_gain, kalman_gain) * measurement_variance
        )
        estimates[index] = state[0]
        velocities[index] = state[1]

    result = pd.DataFrame(
        {"Date": dates, "Kalman": estimates, "Velocity": velocities}
    )
    if source_signature is not None:
        KALMAN_CACHE["series"][cache_key] = result
    return result


def split_signed_series(
    dates: pd.Series,
    values: pd.Series,
) -> tuple[list[Any], list[Any], list[Any], list[Any]]:
    positive_x: list[Any] = []
    positive_y: list[Any] = []
    negative_x: list[Any] = []
    negative_y: list[Any] = []
    points = list(zip(pd.to_datetime(dates), pd.to_numeric(values, errors="coerce")))

    def append_point(target_x, target_y, x_value, y_value) -> None:
        if target_x and target_x[-1] == x_value:
            target_y[-1] = y_value
        else:
            target_x.append(x_value)
            target_y.append(y_value)

    def close_segment(target_x, target_y) -> None:
        if target_x and target_x[-1] is not None:
            target_x.append(None)
            target_y.append(None)

    previous = None
    previous_sign = None
    for x_value, y_value in points:
        if pd.isna(y_value):
            close_segment(positive_x, positive_y)
            close_segment(negative_x, negative_y)
            previous = None
            previous_sign = None
            continue

        y_value = float(y_value)
        current_sign = 1 if y_value > 0 else -1 if y_value < 0 else previous_sign or 1
        if previous is None:
            target_x, target_y = (
                (positive_x, positive_y) if current_sign > 0 else (negative_x, negative_y)
            )
            append_point(target_x, target_y, x_value, y_value)
        else:
            previous_x, previous_y = previous
            if current_sign == previous_sign:
                target_x, target_y = (
                    (positive_x, positive_y) if current_sign > 0 else (negative_x, negative_y)
                )
                append_point(target_x, target_y, x_value, y_value)
            else:
                fraction = abs(previous_y) / (abs(previous_y) + abs(y_value))
                crossing = previous_x + (x_value - previous_x) * fraction
                old_x, old_y = (
                    (positive_x, positive_y) if previous_sign > 0 else (negative_x, negative_y)
                )
                new_x, new_y = (
                    (positive_x, positive_y) if current_sign > 0 else (negative_x, negative_y)
                )
                append_point(old_x, old_y, crossing, 0.0)
                close_segment(old_x, old_y)
                append_point(new_x, new_y, crossing, 0.0)
                append_point(new_x, new_y, x_value, y_value)

        previous = (x_value, y_value)
        previous_sign = current_sign

    return positive_x, positive_y, negative_x, negative_y


def split_series_by_trend(
    dates: pd.Series,
    values: pd.Series,
    trend: pd.Series | None = None,
) -> tuple[list[Any], list[Any], list[Any], list[Any]]:
    trend_up_x: list[Any] = []
    trend_up_y: list[Any] = []
    trend_down_x: list[Any] = []
    trend_down_y: list[Any] = []
    date_values = pd.to_datetime(dates).tolist()
    numeric_values = pd.to_numeric(values, errors="coerce").tolist()
    trend_values = (
        pd.to_numeric(trend, errors="coerce").tolist()
        if trend is not None
        else pd.Series(numeric_values).diff().tolist()
    )
    previous_sign = None

    for index in range(1, len(date_values)):
        previous_value = numeric_values[index - 1]
        current_value = numeric_values[index]
        if pd.isna(previous_value) or pd.isna(current_value):
            previous_sign = None
            continue

        trend_value = trend_values[index]
        current_sign = (
            1
            if not pd.isna(trend_value) and trend_value > 0
            else -1
            if not pd.isna(trend_value) and trend_value < 0
            else previous_sign or 1
        )
        target_x, target_y = (
            (trend_up_x, trend_up_y)
            if current_sign > 0
            else (trend_down_x, trend_down_y)
        )
        if current_sign != previous_sign:
            if target_x and target_x[-1] is not None:
                target_x.append(None)
                target_y.append(None)
            target_x.append(date_values[index - 1])
            target_y.append(float(previous_value))
        target_x.append(date_values[index])
        target_y.append(float(current_value))
        previous_sign = current_sign

    return trend_up_x, trend_up_y, trend_down_x, trend_down_y


def padded_axis_range(values: pd.Series, include_zero: bool = False) -> list[float] | None:
    numeric = pd.to_numeric(values, errors="coerce")
    numeric = numeric[np.isfinite(numeric)].dropna()
    if numeric.empty:
        return None
    minimum = float(numeric.min())
    maximum = float(numeric.max())
    if include_zero:
        maximum_absolute = max(abs(minimum), abs(maximum))
        limit = maximum_absolute * 1.07 if maximum_absolute > 0 else 1.0
        return [-limit, limit]
    if maximum == minimum:
        padding = max(abs(maximum) * 0.05, 1.0)
    else:
        padding = (maximum - minimum) * 0.07
    return [minimum - padding, maximum + padding]


def trading_day_rangebreaks(df: pd.DataFrame) -> list[dict[str, Any]]:
    """Hide weekends and weekday market holidays on Plotly date axes."""
    dates = pd.to_datetime(df["Date"], errors="coerce").dropna()
    if dates.empty:
        return [{"bounds": ["sat", "mon"]}]

    first_day = dates.min().normalize()
    last_day = dates.max().normalize()
    expected_weekdays = pd.date_range(first_day, last_day, freq="B")
    available_days = pd.DatetimeIndex(dates.dt.normalize().unique())
    missing_weekdays = expected_weekdays.difference(available_days)

    breaks: list[dict[str, Any]] = [{"bounds": ["sat", "mon"]}]
    if len(missing_weekdays):
        breaks.append(
            {
                "values": missing_weekdays.strftime("%Y-%m-%d").tolist(),
                "dvalue": 86_400_000,
            }
        )
    return breaks


def build_figure(
    df: pd.DataFrame,
    settings: dict[str, Any],
    trades: pd.DataFrame,
    visible_range: tuple[pd.Timestamp, pd.Timestamp] | None = None,
    instrument_name: str = INSTRUMENT_NAME,
) -> go.Figure:
    window = int(settings["bollinger_window"])
    std_mult = float(settings["bollinger_std"])
    chart_df = (
        cached_bollinger_series(df, window, std_mult)
        if settings["show_bollinger"]
        else df
    )

    full_min_date = pd.Timestamp(chart_df["Date"].min())
    full_max_date = pd.Timestamp(chart_df["Date"].max())
    if visible_range is None:
        min_date = full_min_date
        max_date = full_max_date
    else:
        min_date, max_date = sorted(pd.to_datetime(visible_range))
    plot_df = chart_df[
        (chart_df["Date"] >= min_date) & (chart_df["Date"] <= max_date)
    ].copy()

    active_regressions = []
    regression_color = "#7c3aed"
    regression_dash_style = "dash"
    regression_line_width = 2.1
    regression_labels = {
        182: "6 Monate",
        365: "1 Jahr",
        730: "2 Jahre",
        1825: "5 Jahre",
    }
    if settings["show_regression"]:
        for window_days in settings["regression_windows"]:
            label = regression_labels.get(window_days, f"{window_days} Tage")
            reg = cached_linear_regression_series(df, window_days)
            if not reg.empty:
                active_regressions.append(
                    (label, regression_color, regression_dash_style, reg)
                )

    show_bollinger_difference = bool(settings["show_bollinger"])
    show_kalman = bool(settings["show_kalman"])
    kalman = (
        cached_kalman_filter_series(
            df,
            settings["kalman_process_variance"],
            settings["kalman_measurement_variance"],
        )
        if show_kalman
        else pd.DataFrame(columns=["Date", "Kalman", "Velocity"])
    )
    difference_count = (
        len(active_regressions)
        + int(show_bollinger_difference)
        + int(show_kalman)
    )
    spread_count = difference_count
    row_count = 1 + spread_count
    main_panel_height = 420
    difference_panel_height = 105
    row_heights = [main_panel_height] + [difference_panel_height] * spread_count
    figure_height = (
        main_panel_height
        + difference_panel_height * spread_count
        + 114
    )

    fig = make_subplots(
        rows=row_count,
        cols=1,
        shared_xaxes=True,
        vertical_spacing=0.0,
        row_heights=row_heights,
    )

    queued_traces = []
    trace_rows = []
    trace_columns = []

    def queue_trace(trace, row, col):
        queued_traces.append(trace)
        trace_rows.append(row)
        trace_columns.append(col)

    queue_trace(
        go.Scatter(
            x=plot_df["Date"],
            y=plot_df["Price"],
            mode="lines",
            name=instrument_name,
            line={"width": 2.5, "color": "#0f172a"},
            hovertemplate=f"{instrument_name}: %{{y:.4f}}<extra></extra>",
        ),
        row=1,
        col=1,
    )

    if settings["show_bollinger"]:
        queue_trace(
            go.Scatter(
                x=plot_df["Date"],
                y=plot_df["BB_Upper"],
                mode="lines",
                name="Bollinger Upper",
                hoverinfo="skip",
                line={"width": 1.2, "color": "#60a5fa", "dash": "dot"},
            ),
            row=1,
            col=1,
        )
        queue_trace(
            go.Scatter(
                x=plot_df["Date"],
                y=plot_df["BB_Lower"],
                mode="lines",
                name="Bollinger Lower",
                hoverinfo="skip",
                line={"width": 1.2, "color": "#60a5fa", "dash": "dot"},
                fill="tonexty",
                fillcolor="rgba(59, 130, 246, 0.08)",
            ),
            row=1,
            col=1,
        )
        queue_trace(
            go.Scatter(
                x=plot_df["Date"],
                y=plot_df["BB_Middle"],
                mode="lines",
                name="Bollinger Mittelwert",
                hoverinfo="skip",
                line={"width": 1.4, "color": "#2563eb"},
            ),
            row=1,
            col=1,
        )

    regression_endpoints = []
    for label, color, dash_style, reg in active_regressions:
        visible_reg = reg[
            (reg["Date"] >= min_date) & (reg["Date"] <= max_date)
        ].dropna(subset=["Regression"])
        if visible_reg.empty:
            continue
        queue_trace(
            go.Scatter(
                x=visible_reg["Date"],
                y=visible_reg["Regression"],
                mode="lines",
                name=f"Regression {label}",
                hoverinfo="skip",
                line={"width": regression_line_width, "dash": dash_style, "color": color},
            ),
            row=1,
            col=1,
        )
        last_point = visible_reg.iloc[-1]
        regression_endpoints.append(
            (label, color, last_point["Date"], float(last_point["Regression"]))
        )

    if show_kalman:
        visible_kalman = kalman[
            (kalman["Date"] >= min_date) & (kalman["Date"] <= max_date)
        ]
        up_x, up_y, down_x, down_y = split_series_by_trend(
            visible_kalman["Date"],
            visible_kalman["Kalman"],
        )
        for trace_name, trace_x, trace_y, trace_color in (
            ("Kalman 2D steigend", up_x, up_y, "#16a34a"),
            ("Kalman 2D fallend", down_x, down_y, "#dc2626"),
        ):
            if trace_x:
                queue_trace(
                    go.Scatter(
                        x=trace_x,
                        y=trace_y,
                        mode="lines",
                        name=trace_name,
                        hoverinfo="skip",
                        line={"width": 2.0, "color": trace_color},
                        showlegend=False,
                        connectgaps=False,
                    ),
                    row=1,
                    col=1,
                )

    if not trades.empty:
        entry_trades = trades.dropna(subset=["Entry_Date", "Entry_Price"])
        entry_trades = entry_trades[
            (entry_trades["Entry_Date"] >= min_date)
            & (entry_trades["Entry_Date"] <= max_date)
        ]
        if not entry_trades.empty:
            queue_trace(
                go.Scatter(
                    x=entry_trades["Entry_Date"],
                    y=entry_trades["Entry_Price"],
                    mode="markers",
                    name="Kauf",
                    hoverinfo="skip",
                    marker={"symbol": "triangle-up", "size": 11, "color": "#16a34a", "line": {"width": 1.2, "color": "white"}},
                    customdata=np.stack(
                        [
                            entry_trades["Trade_ID"].astype(str),
                            entry_trades["Direction"].astype(str),
                        ],
                        axis=-1,
                    ),
                ),
                row=1,
                col=1,
            )

        exit_trades = trades.dropna(subset=["Exit_Date", "Exit_Price"])
        exit_trades = exit_trades[
            (exit_trades["Exit_Date"] >= min_date)
            & (exit_trades["Exit_Date"] <= max_date)
        ]
        if not exit_trades.empty:
            queue_trace(
                go.Scatter(
                    x=exit_trades["Exit_Date"],
                    y=exit_trades["Exit_Price"],
                    mode="markers",
                    name="Verkauf",
                    hoverinfo="skip",
                    marker={"symbol": "triangle-down", "size": 11, "color": "#dc2626", "line": {"width": 1.2, "color": "white"}},
                    customdata=exit_trades["Trade_ID"].astype(str),
                ),
                row=1,
                col=1,
            )

    difference_panels = []
    difference_series = []
    for spread_row, (label, color, _dash_style, reg) in enumerate(
        active_regressions, start=2
    ):
        visible_reg = reg[
            (reg["Date"] >= min_date) & (reg["Date"] <= max_date)
        ]
        spread_df = visible_reg.merge(plot_df[["Date", "Price"]], on="Date", how="left")
        spread_df["Spread"] = spread_df["Price"] - spread_df["Regression"]
        difference_panels.append((spread_row, label, color, spread_df))
        difference_series.append((spread_row, label, spread_df))

    if show_bollinger_difference:
        bollinger_values = plot_df[
            ["Date", "Price", "BB_Upper", "BB_Lower"]
        ].dropna().copy()
        upper_difference = bollinger_values[["Date"]].copy()
        upper_difference["Spread"] = (
            bollinger_values["Price"] - bollinger_values["BB_Upper"]
        )
        lower_difference = bollinger_values[["Date"]].copy()
        lower_difference["Spread"] = (
            bollinger_values["Price"] - bollinger_values["BB_Lower"]
        )
        bollinger_range = pd.concat(
            [upper_difference, lower_difference], ignore_index=True
        )
        bollinger_row = 2 + len(active_regressions)
        difference_panels.append(
            (bollinger_row, "Bollinger Bands", "#2563eb", bollinger_range)
        )
        difference_series.extend(
            [
                (bollinger_row, "Kurs - Upper Band", upper_difference),
                (bollinger_row, "Kurs - Lower Band", lower_difference),
            ]
        )

    if show_kalman:
        daily_kalman = kalman[["Date", "Kalman"]].dropna().copy()
        daily_kalman["Trading_Day"] = daily_kalman["Date"].dt.normalize()
        daily_kalman = (
            daily_kalman.sort_values("Date")
            .groupby("Trading_Day", as_index=False)
            .last()
        )
        daily_kalman["Spread"] = daily_kalman["Kalman"].diff()
        kalman_slope = daily_kalman[
            (daily_kalman["Date"] >= min_date)
            & (daily_kalman["Date"] <= max_date)
        ].dropna(subset=["Spread"])
        kalman_row = (
            2 + len(active_regressions) + int(show_bollinger_difference)
        )
        difference_panels.append(
            (kalman_row, "Kalman-Steigung zum Vortag", "#db2777", kalman_slope)
        )
        if not kalman_slope.empty:
            queue_trace(
                go.Bar(
                    x=kalman_slope["Date"],
                    y=kalman_slope["Spread"],
                    name="Kalman-Steigung zum Vortag",
                    marker={
                        "color": np.where(
                            kalman_slope["Spread"] >= 0,
                            "#16a34a",
                            "#dc2626",
                        )
                    },
                    hovertemplate="Steigung zum Vortag: %{y:.4f}<extra></extra>",
                    showlegend=False,
                ),
                row=kalman_row,
                col=1,
            )

    for spread_row, label, spread_df in difference_series:
        positive_x, positive_y, negative_x, negative_y = split_signed_series(
            spread_df["Date"], spread_df["Spread"]
        )
        if positive_x:
            queue_trace(
                go.Scatter(
                    x=positive_x,
                    y=positive_y,
                    mode="lines",
                    line={"width": 1.9, "color": "#16a34a"},
                    fill="tozeroy",
                    fillcolor="rgba(22, 163, 74, 0.11)",
                    name=f"{label} positiv",
                    hoverinfo="none",
                    showlegend=False,
                    connectgaps=False,
                ),
                row=spread_row,
                col=1,
            )
        if negative_x:
            queue_trace(
                go.Scatter(
                    x=negative_x,
                    y=negative_y,
                    mode="lines",
                    line={"width": 1.9, "color": "#dc2626"},
                    fill="tozeroy",
                    fillcolor="rgba(220, 38, 38, 0.10)",
                    name=f"{label} negativ",
                    hoverinfo="none",
                    showlegend=False,
                    connectgaps=False,
                ),
                row=spread_row,
                col=1,
            )

    if queued_traces:
        fig.add_traces(queued_traces, rows=trace_rows, cols=trace_columns)

    main_values = []
    for trace in fig.data:
        if trace.yaxis not in (None, "y") or trace.y is None:
            continue
        numeric_values = pd.to_numeric(pd.Series(trace.y), errors="coerce").to_numpy(
            dtype=float
        )
        finite_values = numeric_values[np.isfinite(numeric_values)]
        if finite_values.size:
            main_values.extend(finite_values.tolist())

    yaxis_range = padded_axis_range(pd.Series(main_values))

    fig.update_layout(
        template="plotly_white",
        hovermode="x unified",
        hoversubplots="axis",
        hoverdistance=-1,
        showlegend=False,
        height=figure_height,
        autosize=True,
        paper_bgcolor="#ffffff",
        plot_bgcolor="#ffffff",
        font={"family": "Arial, sans-serif", "color": "#0f172a"},
        margin={"l": 48, "r": 88, "t": 72, "b": 42},
        hoverlabel={
            "bgcolor": "#0f172a",
            "bordercolor": "#0f172a",
            "font": {"color": "#ffffff", "size": 12},
        },
        legend={
            "orientation": "h",
            "y": 1.07,
            "x": 1.0,
            "xanchor": "right",
            "yanchor": "bottom",
            "bgcolor": "rgba(255,255,255,0)",
            "itemsizing": "constant",
            "font": {"size": 11},
        },
    )

    fig.update_xaxes(
        title_text="",
        showgrid=False,
        showline=False,
        ticks="",
        showticklabels=False,
        tickfont={"size": 10, "color": "#64748b"},
        hoverformat="%d.%m.%Y",
        rangeslider_visible=False,
        automargin=True,
        showspikes=False,
        range=[min_date, max_date],
        rangebreaks=trading_day_rangebreaks(df),
    )
    fig.update_yaxes(
        title_text=None,
        type="linear",
        range=yaxis_range,
        showgrid=False,
        showline=False,
        zeroline=False,
        tickfont={"size": 10, "color": "#64748b"},
        tickformat=".3f",
        automargin=True,
        row=1,
        col=1,
    )
    for label, color, end_date, end_value in regression_endpoints:
        fig.add_annotation(
            x=end_date,
            y=end_value,
            xref="x",
            yref="y",
            text=label,
            showarrow=False,
            xanchor="left",
            yanchor="middle",
            xshift=7,
            font={"family": "Arial, sans-serif", "size": 10, "color": color},
            bgcolor="rgba(255, 255, 255, 0.88)",
            borderpad=2,
        )
    for spread_row, label, color, spread_df in difference_panels:
        visible_difference = spread_df[
            (spread_df["Date"] >= min_date) & (spread_df["Date"] <= max_date)
        ]
        difference_range = padded_axis_range(
            visible_difference["Spread"], include_zero=True
        )
        fig.update_yaxes(
            title_text=None,
            range=difference_range,
            showgrid=False,
            showline=False,
            zeroline=False,
            showticklabels=False,
            ticks="",
            automargin=True,
            row=spread_row,
            col=1,
        )
        fig.add_hline(
            y=0,
            line={"color": "#111827", "width": 0.55},
            layer="above",
            row=spread_row,
            col=1,
        )
        panel_label = (
            label
            if label in {"Bollinger Bands", "Kalman-Steigung zum Vortag"}
            else f"Regression {label}"
        )
        fig.add_annotation(
            x=0.006,
            y=0,
            xref="paper",
            yref=f"y{spread_row}",
            text=panel_label,
            showarrow=False,
            xanchor="left",
            yanchor="bottom",
            yshift=3,
            font={"family": "Arial, sans-serif", "size": 10, "color": color},
            opacity=0.52,
        )
    fig.update_xaxes(
        ticks="",
        ticklen=0,
        showticklabels=True,
        tickfont={"size": 10, "color": "#64748b"},
        row=row_count,
        col=1,
    )
    return fig


def trade_summary(trades: pd.DataFrame) -> dict[str, Any]:
    if trades.empty:
        return {
            "closed_trades": 0,
            "open_trades": 0,
            "hit_rate": np.nan,
            "avg_return": np.nan,
            "median_return": np.nan,
            "total_pnl": 0.0,
            "profit_factor": np.nan,
            "avg_holding_days": np.nan,
            "best_trade": np.nan,
            "worst_trade": np.nan,
        }

    closed = trades[trades["Status"].astype(str).str.lower() == "closed"].copy()
    open_count = int((trades["Status"].astype(str).str.lower() == "open").sum())

    if closed.empty:
        return {
            "closed_trades": 0,
            "open_trades": open_count,
            "hit_rate": np.nan,
            "avg_return": np.nan,
            "median_return": np.nan,
            "total_pnl": 0.0,
            "profit_factor": np.nan,
            "avg_holding_days": np.nan,
            "best_trade": np.nan,
            "worst_trade": np.nan,
        }

    returns = closed["Net_Return_Pct"].dropna()
    pnl = closed["Net_PnL"].dropna()
    gains = pnl[pnl > 0].sum()
    losses = abs(pnl[pnl < 0].sum())

    return {
        "closed_trades": int(len(closed)),
        "open_trades": open_count,
        "hit_rate": float((returns > 0).mean() * 100) if len(returns) else np.nan,
        "avg_return": float(returns.mean()) if len(returns) else np.nan,
        "median_return": float(returns.median()) if len(returns) else np.nan,
        "total_pnl": float(pnl.sum()) if len(pnl) else 0.0,
        "profit_factor": float(gains / losses) if losses > 0 else np.nan,
        "avg_holding_days": float(closed["Holding_Days"].mean()),
        "best_trade": float(returns.max()) if len(returns) else np.nan,
        "worst_trade": float(returns.min()) if len(returns) else np.nan,
    }


def fmt_pct(value: Any) -> str:
    return "–" if pd.isna(value) else f"{value:.2f} %"


def fmt_num(value: Any) -> str:
    return "–" if pd.isna(value) else f"{value:,.2f}"


def metric_tone(value: Any) -> str:
    if pd.isna(value) or float(value) == 0:
        return "neutral"
    return "positive" if float(value) > 0 else "negative"


def summary_hero(
    title: str, value: str, subtitle: str, tone: str = "neutral"
) -> html.Div:
    return html.Div(
        [
            html.Div(title, className="summary-hero-label"),
            html.Div(value, className="summary-hero-value"),
            html.Div(subtitle, className="summary-hero-subtitle"),
        ],
        className=f"summary-hero summary-tone-{tone}",
    )


def summary_stat(title: str, value: str, tone: str = "neutral") -> html.Div:
    return html.Div(
        [
            html.Span(title, className="summary-stat-label"),
            html.Span(value, className=f"summary-stat-value summary-tone-{tone}"),
        ],
        className="summary-stat",
    )


def visible_range_from_relayout(
    relayout_data: dict[str, Any] | None,
) -> tuple[pd.Timestamp, pd.Timestamp] | None:
    if not relayout_data:
        return None
    if any(
        key.endswith(".autorange") and value is True
        for key, value in relayout_data.items()
        if key.startswith("xaxis")
    ):
        return None

    axis_names = ["xaxis"] + [f"xaxis{index}" for index in range(2, 9)]
    for axis_name in axis_names:
        combined = relayout_data.get(f"{axis_name}.range")
        if isinstance(combined, (list, tuple)) and len(combined) == 2:
            try:
                return pd.Timestamp(combined[0]), pd.Timestamp(combined[1])
            except (TypeError, ValueError):
                pass
        start = relayout_data.get(f"{axis_name}.range[0]")
        end = relayout_data.get(f"{axis_name}.range[1]")
        if start is not None and end is not None:
            try:
                return pd.Timestamp(start), pd.Timestamp(end)
            except (TypeError, ValueError):
                pass
    return None


def visible_range_from_control(
    market_df: pd.DataFrame,
    selection: str | None,
) -> tuple[pd.Timestamp, pd.Timestamp] | None:
    if not selection or selection == "max":
        return None

    data_start = pd.Timestamp(market_df["Date"].min())
    data_end = pd.Timestamp(market_df["Date"].max())
    offsets = {
        "1d": pd.Timedelta(days=1),
        "5d": pd.Timedelta(days=5),
        "1m": pd.DateOffset(months=1),
        "2m": pd.DateOffset(months=2),
    }
    if selection in offsets:
        return max(data_start, data_end - offsets[selection]), data_end
    return None


ensure_setup_structure()
initial_settings = load_settings()

app = Dash(__name__)
app.title = f"{INSTRUMENT_NAME} Analyse"

app.layout = html.Div(
    [
        html.Div(
            [
                html.H2(
                    "MSCI World ETF Analyse",
                    style={"margin": "0", "fontSize": "28px", "fontWeight": "700", "letterSpacing": "-0.02em"},
                ),
                dcc.Dropdown(
                    id="instrument-selector",
                    options=[
                        {"label": config["name"], "value": key}
                        for key, config in INSTRUMENTS.items()
                    ],
                    value=DEFAULT_INSTRUMENT_KEY,
                    clearable=False,
                    style={"minWidth": "340px", "marginLeft": "auto"},
                ),
                dcc.ConfirmDialogProvider(
                    html.Button(
                        "Kursdaten neu laden",
                        id="reload-market-data-button",
                        n_clicks=0,
                        className="primary-action-button",
                        style={"flexShrink": "0", "whiteSpace": "nowrap"},
                    ),
                    id="reload-market-data-confirm",
                    message="Sollen Daily-, 5-Minuten- und Last-Price-Daten jetzt von Yahoo Finance neu geladen werden?",
                ),
            ],
            style={
                "display": "flex",
                "alignItems": "center",
                "justifyContent": "space-between",
                "gap": "16px",
                "marginBottom": "8px",
            },
        ),
        html.Div(
            id="instrument-meta",
            children=f"ISIN {INSTRUMENT_ISIN} · Yahoo {YAHOO_TICKER} · Trading Tools: Tagesdaten (MAX, Adj Close) · Trading Analytics: 5-Minuten-Daten",
            style={"marginBottom": "12px", "color": "#475569", "fontSize": "13px"},
        ),
        html.Div(id="market-reload-message", className="message"),
        dcc.Store(id="market-data-refresh-token", data=0),
        dcc.Tabs(
            id="main-tabs",
            value="trading-tools",
            className="main-tabs",
            children=[
                dcc.Tab(
                    label="Trading Tools",
                    value="trading-tools",
                    className="main-tab",
                    selected_className="main-tab--selected",
                    children=[
        html.Div(
            [
                html.Div(
                    [
                        html.Div(
                            [
                                dcc.Graph(
                                    id="main-chart",
                                    style={"width": "100%"},
                                    config={
                                        "displaylogo": False,
                                        "responsive": True,
                                        "modeBarButtonsToAdd": [
                                            "drawline",
                                            "drawopenpath",
                                            "eraseshape",
                                            "resetScale2d",
                                        ],
                                    },
                                ),
                                html.Div(
                                    [
                                        dcc.RadioItems(
                                            id="time-range-selector",
                                            options=[
                                                {"label": "1T", "value": "1d"},
                                                {"label": "5T", "value": "5d"},
                                                {"label": "1M", "value": "1m"},
                                                {"label": "2M", "value": "2m"},
                                                {"label": "MAX", "value": "max"},
                                            ],
                                            value="1m",
                                            inline=True,
                                            className="time-range-buttons",
                                        ),
                                    ],
                                    className="time-range-card",
                                ),
                            ],
                            className="panel chart-panel",
                        ),
                    ],
                    className="chart-column",
                ),
                html.Div(
                    [
                        html.Div(
                            [
                                dcc.Checklist(
                                    id="regression-toggle",
                                    options=[{"label": "Lineare Regression", "value": "show"}],
                                    value=["show"] if any(
                                        initial_settings[key]
                                        for key in (
                                            "show_regression_6m",
                                            "show_regression_1y",
                                            "show_regression_2y",
                                            "show_regression_5y",
                                            "show_regression_max",
                                        )
                                    ) else [],
                                    inline=True,
                                    className="regression-button",
                                ),
                                html.Div(
                                    [
                                        html.Div(
                                            [
                                                html.Label("Kurzfristig (Tage)", className="compact-label"),
                                                dcc.Input(id="reg-6m-days", type="text", debounce=True, value=str(initial_settings["regression_6m_days"])),
                                            ],
                                            className="setting-field",
                                        ),
                                        html.Div(
                                            [
                                                html.Label("Mittelfristig (Tage)", className="compact-label"),
                                                dcc.Input(id="reg-1y-days", type="text", debounce=True, value=str(initial_settings["regression_1y_days"])),
                                            ],
                                            className="setting-field",
                                        ),
                                        html.Div(
                                            [
                                                html.Label("Langfristig (Tage)", className="compact-label"),
                                                dcc.Input(id="reg-2y-days", type="text", debounce=True, value=str(initial_settings["regression_5y_days"])),
                                            ],
                                            className="setting-field",
                                        ),
                                    ],
                                    className="regression-settings",
                                ),
                            ],
                            className="indicator-card regression-card",
                        ),
                        html.Div(
                            [
                                dcc.Checklist(
                                    id="bollinger-toggle",
                                    options=[{"label": "Bollinger Bands", "value": "show"}],
                                    value=["show"] if initial_settings["show_bollinger"] else [],
                                    inline=True,
                                    className="bollinger-button",
                                ),
                                html.Div(
                                    [
                                        html.Div(
                                            [
                                                html.Label("Periode (Tage)", className="compact-label"),
                                                dcc.Input(id="bollinger-window", type="text", debounce=True, value=str(initial_settings["bollinger_window"])),
                                            ],
                                            className="setting-field",
                                        ),
                                        html.Div(
                                            [
                                                html.Label("Std.-Abweichung", className="compact-label"),
                                                dcc.Input(id="bollinger-std", type="text", debounce=True, value=str(initial_settings["bollinger_std"])),
                                            ],
                                            className="setting-field",
                                        ),
                                    ],
                                    className="bollinger-settings",
                                ),
                            ],
                            className="indicator-card bollinger-card",
                        ),
                        html.Div(
                            [
                                dcc.Checklist(
                                    id="kalman-toggle",
                                    options=[{"label": "Kalman-Filter 2D", "value": "show"}],
                                    value=["show"] if initial_settings["show_kalman"] else [],
                                    inline=True,
                                    className="kalman-button",
                                ),
                                html.Div(
                                    [
                                        html.Div(
                                            [
                                                html.Div(
                                                    [
                                                        html.Label("Prozess Q", className="compact-label"),
                                                        html.Span(
                                                            [
                                                                html.Span("i", className="parameter-info-glyph"),
                                                                html.Span(
                                                                    "Höher = reagiert schneller auf Richtungswechsel",
                                                                    className="parameter-tooltip",
                                                                ),
                                                            ],
                                                            className="parameter-info",
                                                            tabIndex=0,
                                                        ),
                                                    ],
                                                    className="parameter-label-row",
                                                ),
                                                dcc.Input(
                                                    id="kalman-process-variance",
                                                    type="text",
                                                    debounce=True,
                                                    value=str(initial_settings["kalman_process_variance"]),
                                                ),
                                            ],
                                            className="setting-field",
                                        ),
                                        html.Div(
                                            [
                                                html.Div(
                                                    [
                                                        html.Label("Messung R", className="compact-label"),
                                                        html.Span(
                                                            [
                                                                html.Span("i", className="parameter-info-glyph"),
                                                                html.Span(
                                                                    "Höher = glättet Kursrauschen stärker",
                                                                    className="parameter-tooltip",
                                                                ),
                                                            ],
                                                            className="parameter-info",
                                                            tabIndex=0,
                                                        ),
                                                    ],
                                                    className="parameter-label-row",
                                                ),
                                                dcc.Input(
                                                    id="kalman-measurement-variance",
                                                    type="text",
                                                    debounce=True,
                                                    value=str(initial_settings["kalman_measurement_variance"]),
                                                ),
                                            ],
                                            className="setting-field",
                                        ),
                                    ],
                                    className="kalman-settings",
                                ),
                            ],
                            className="indicator-card kalman-card",
                        ),
                        html.Button(
                            "Parameter als Standard speichern",
                            id="save-settings",
                            n_clicks=0,
                            className="primary-action-button",
                            style={"width": "100%"},
                        ),
                        html.Div(id="settings-message", className="message"),
                    ],
                    className="indicator-sidebar",
                ),
            ],
            className="chart-workspace",
        ),
                    ],
                ),
                dcc.Tab(
                    label="Trading Analytics",
                    value="trading-analytics",
                    className="main-tab",
                    selected_className="main-tab--selected",
                    children=[
                        html.Div(
                            [
                                html.Div(
                                    [
                                        html.Div(
                                            [
                                                html.Div("Kurs, Depot & Vergleich", className="analytics-panel-title"),
                                                dcc.RadioItems(
                                                    id="analytics-time-range",
                                                    options=[
                                                        {"label": "1J", "value": "1y"},
                                                        {"label": "2J", "value": "2y"},
                                                        {"label": "5J", "value": "5y"},
                                                        {"label": "MAX", "value": "max"},
                                                    ],
                                                    value="max",
                                                    inline=True,
                                                    className="analytics-time-range",
                                                ),
                                            ],
                                            className="analytics-chart-header",
                                        ),
                                        dcc.Graph(id="analytics-chart", style={"width": "100%", "height": "100%"}, config={"displaylogo": False, "responsive": True}),
                                    ],
                                    className="analytics-panel analytics-chart-panel",
                                ),
                                html.Div(
                                    [
                                        html.Div(
                                            [
                                                html.Div("Auswertung", className="analytics-panel-title"),
                                                html.Div(id="trade-metrics", className="metric-grid analytics-metrics"),
                                            ],
                                            className="analytics-panel analytics-summary-panel",
                                        ),
                                        html.Div(
                                            [
                                                html.Div(
                                                    [
                                                        html.Div("Einzelne Trades", className="analytics-panel-title"),
                                                        html.Div(
                                                            [
                                                                dcc.ConfirmDialogProvider(
                                                                    html.Button("Trade löschen", id="delete-trade", n_clicks=0, disabled=True, className="danger-action-button"),
                                                                    id="delete-trade-confirm",
                                                                    message="Soll der ausgewählte Trade wirklich gelöscht werden?",
                                                                ),
                                                                html.Button("Trade erfassen", id="open-trade-modal", n_clicks=0, className="primary-action-button"),
                                                            ],
                                                            className="trade-header-actions",
                                                        ),
                                                    ],
                                                    className="analytics-trades-header",
                                                ),
                                                html.Div(
                                                    dash_table.DataTable(
                                                        id="trade-table",
                                                        page_size=6,
                                                        sort_action="native",
                                                        style_table={"overflowX": "auto", "maxHeight": "265px", "overflowY": "auto"},
                                                        style_cell={"fontFamily": "Arial, sans-serif", "fontSize": 11, "padding": "7px", "textAlign": "right", "minWidth": "80px", "maxWidth": "150px", "whiteSpace": "nowrap"},
                                                        style_header={"fontWeight": "700", "backgroundColor": "#f8fafc", "borderColor": "#e2e8f0"},
                                                        style_data_conditional=[
                                                            {"if": {"state": "active"}, "backgroundColor": "#eef2ff", "border": "1px solid #818cf8"},
                                                        ],
                                                    ),
                                                    className="analytics-table-wrap",
                                                ),
                                            ],
                                            className="analytics-panel analytics-trades-panel",
                                        ),
                                    ],
                                    className="analytics-side-column",
                                ),
                            ],
                            className="analytics-grid",
                        ),
                        dcc.Store(id="refresh-token", data=0),
                        html.Div(
                            html.Div(
                                [
                                    html.Div(
                                        [
                                            html.Div("Trade erfassen", className="trade-modal-title"),
                                            html.Button("×", id="close-trade-modal", n_clicks=0, className="modal-close-button", title="Schließen"),
                                        ],
                                        className="trade-modal-header",
                                    ),
                                    html.Div(
                                        [
                                            html.Div([html.Label("Entry-Datum"), dcc.DatePickerSingle(id="entry-date", display_format="DD.MM.YYYY", date=date.today())], className="form-field"),
                                            html.Div([html.Label("Exit-Datum"), dcc.DatePickerSingle(id="exit-date", display_format="DD.MM.YYYY", clearable=True)], className="form-field"),
                                            html.Div([html.Label("Entry-Preis (leer = Kurswert)"), dcc.Input(id="entry-price", type="number", min=0)], className="form-field"),
                                            html.Div([html.Label("Exit-Preis (leer = Kurswert)"), dcc.Input(id="exit-price", type="number", min=0)], className="form-field"),
                                            html.Div([html.Label("Gebühren gesamt"), dcc.Input(id="fees", type="number", min=0, value=0.0)], className="form-field"),
                                            html.Div([html.Label("Notiz"), dcc.Input(id="notes", type="text", placeholder="Signal, These oder Kommentar")], className="form-field"),
                                        ],
                                        className="trade-form",
                                    ),
                                    html.Div(id="trade-message", className="message"),
                                    html.Div(
                                        [
                                            html.Button("Abbrechen", id="cancel-trade-modal", n_clicks=0, className="secondary-action-button"),
                                            html.Button("Trade speichern", id="save-trade", n_clicks=0, className="primary-action-button"),
                                        ],
                                        className="trade-modal-actions",
                                    ),
                                ],
                                className="trade-modal",
                            ),
                            id="trade-modal-overlay",
                            className="trade-modal-overlay",
                            style={"display": "none"},
                        ),
                    ],
                ),
            ],
        ),
        dcc.Store(id="hover-sync-token"),
    ],
    style={
        "fontFamily": "Arial, sans-serif",
        "width": "100%",
        "maxWidth": "none",
        "margin": "0 auto",
        "padding": "clamp(8px, 1.5vw, 24px)",
        "boxSizing": "border-box",
        "background": "#ffffff",
    },
)


app.index_string = """
<!DOCTYPE html>
<html>
    <head>
        {%metas%}
        <title>{%title%}</title>
        {%favicon%}
        {%css%}
        <style>
            body {
                margin: 0;
                background: #ffffff;
                color: #0f172a;
                font-family: Arial, sans-serif;
            }
            body * {
                font-family: Arial, sans-serif !important;
            }
            .main-tabs {
                width: 100%;
            }
            .main-tabs .tab-container {
                display: grid !important;
                grid-template-columns: repeat(2, minmax(180px, 1fr));
                gap: 8px;
                padding: 7px;
                border: 1px solid #cbd5e1;
                border-radius: 15px;
                background: #eef2ff;
                box-shadow: 0 8px 22px rgba(15, 23, 42, 0.08);
            }
            .main-tab {
                display: flex !important;
                align-items: center;
                justify-content: center;
                min-height: 48px;
                padding: 0 22px !important;
                border: 1px solid #cbd5e1 !important;
                border-radius: 10px;
                background: #ffffff !important;
                color: #334155 !important;
                font-size: 14px;
                font-weight: 800;
                letter-spacing: 0.02em;
            }
            .main-tab--selected {
                border-color: #4f46e5 !important;
                background: linear-gradient(135deg, #4f46e5, #7c3aed) !important;
                color: #ffffff !important;
                box-shadow: 0 7px 16px rgba(79, 70, 229, 0.24);
            }
            .control-grid {
                display: grid;
                grid-template-columns: repeat(auto-fit, minmax(320px, 1fr));
                gap: 18px;
                margin-bottom: 18px;
            }
            .panel {
                border: 1px solid rgba(148, 163, 184, 0.35);
                border-radius: 16px;
                background: rgba(255, 255, 255, 0.9);
                backdrop-filter: blur(8px);
                padding: 18px;
                box-shadow: 0 12px 28px rgba(15, 23, 42, 0.08);
                box-sizing: border-box;
            }
            .chart-panel {
                grid-area: chart;
                position: relative;
                z-index: 1;
                border: 1px solid rgba(148, 163, 184, 0.35);
                border-radius: 20px;
                background: white;
                box-shadow: 0 18px 44px rgba(15, 23, 42, 0.08);
                flex: 1 1 auto;
                min-width: 0;
                padding: 0;
                overflow: hidden;
            }
            .chart-workspace {
                display: grid;
                grid-template-columns: minmax(0, 1fr);
                grid-template-areas:
                    "indicators"
                    "chart";
                align-items: stretch;
                gap: clamp(12px, 1.25vw, 20px);
                width: 100%;
                margin: 14px 0 20px;
            }
            .chart-column {
                display: contents;
            }
            .indicator-sidebar {
                display: grid;
                grid-area: indicators;
                grid-template-columns: repeat(3, minmax(0, 1fr));
                gap: 10px;
                width: 100%;
                min-width: 0;
            }
            .indicator-card {
                padding: 13px;
                border: 1px solid rgba(148, 163, 184, 0.28);
                border-radius: 14px;
                background: rgba(255, 255, 255, 0.86);
                backdrop-filter: blur(10px);
                box-shadow: 0 10px 26px rgba(15, 23, 42, 0.10);
                box-sizing: border-box;
            }
            .regression-card {
                border-top: 3px solid #7c3aed;
            }
            .bollinger-card {
                border-top: 3px solid #2563eb;
            }
            .kalman-card {
                border-top: 3px solid #db2777;
            }
            .indicator-title {
                font-family: Arial, sans-serif !important;
                color: #0f172a;
                font-size: 12px;
                font-weight: 800;
                letter-spacing: 0.12em;
            }
            .indicator-subtitle {
                margin: 5px 0 14px;
                color: #64748b;
                font-size: 12px;
            }
            .side-panel {
                min-height: 0;
            }
            .control-row {
                display: flex;
                justify-content: space-between;
                align-items: center;
                gap: 12px;
                margin: 12px 0;
                padding: 10px 12px;
                min-height: 52px;
                background: #f8fafc;
                border-radius: 10px;
                border: 1px solid rgba(148, 163, 184, 0.2);
                box-sizing: border-box;
            }
            .trade-form {
                display: grid;
                grid-template-columns: repeat(2, minmax(0, 1fr));
                gap: 14px;
                margin-bottom: 14px;
            }
            .analytics-grid {
                display: grid;
                grid-template-columns: repeat(2, minmax(0, 1fr));
                align-items: stretch;
                gap: clamp(14px, 1.5vw, 22px);
                width: 100%;
                margin: 18px 0 22px;
            }
            .analytics-panel {
                display: flex;
                flex-direction: column;
                width: 100%;
                min-width: 0;
                padding: 16px;
                border: 1px solid rgba(148, 163, 184, 0.35);
                border-radius: 18px;
                background: #ffffff;
                box-shadow: 0 14px 34px rgba(15, 23, 42, 0.07);
                box-sizing: border-box;
                overflow: hidden;
            }
            .analytics-chart-panel,
            .analytics-side-column {
                height: 720px;
            }
            .analytics-side-column {
                display: grid;
                grid-template-rows: repeat(2, minmax(0, 1fr));
                gap: clamp(14px, 1.5vw, 22px);
                min-width: 0;
            }
            .analytics-summary-panel,
            .analytics-trades-panel {
                min-height: 0;
            }
            .analytics-chart-header,
            .analytics-trades-header {
                display: flex;
                flex: 0 0 auto;
                align-items: center;
                justify-content: space-between;
                gap: 12px;
                margin-bottom: 12px;
            }
            .analytics-chart-header .analytics-panel-title,
            .analytics-trades-header .analytics-panel-title {
                margin-bottom: 0;
            }
            .analytics-panel-title {
                flex: 0 0 auto;
                margin-bottom: 12px;
                color: #334155;
                font-size: 12px;
                font-weight: 800;
                letter-spacing: 0.08em;
                text-transform: uppercase;
            }
            .analytics-chart-panel #analytics-chart,
            .analytics-chart-panel .js-plotly-plot,
            .analytics-chart-panel .plot-container {
                flex: 1 1 auto;
                width: 100% !important;
                min-height: 0;
            }
            .analytics-table-wrap {
                flex: 1 1 auto;
                min-width: 0;
                min-height: 0;
                overflow: auto;
            }
            .analytics-metrics {
                display: flex;
                flex: 1 1 auto;
                margin: 0;
                min-height: 0;
            }
            .evaluation-table {
                display: grid;
                flex: 1 1 auto;
                grid-template-rows: 40px repeat(5, minmax(38px, 1fr));
                width: 100%;
                height: 100%;
                min-width: 0;
                overflow: hidden;
                border: 1px solid #e2e8f0;
                border-radius: 12px;
                background: #ffffff;
                box-sizing: border-box;
            }
            .evaluation-row {
                display: grid;
                grid-template-columns: minmax(120px, 1.35fr) repeat(3, minmax(78px, 1fr));
                align-items: stretch;
                min-width: 0;
                border-bottom: 1px solid #e2e8f0;
            }
            .evaluation-row:last-child {
                border-bottom: 0;
            }
            .evaluation-row > div {
                display: flex;
                align-items: center;
                justify-content: flex-end;
                min-width: 0;
                padding: 6px 10px;
                border-left: 1px solid rgba(226, 232, 240, 0.76);
                color: #334155;
                font-size: 10px;
                font-variant-numeric: tabular-nums;
                text-align: right;
                box-sizing: border-box;
            }
            .evaluation-row > div:first-child {
                justify-content: flex-start;
                border-left: 0;
                color: #475569;
                font-weight: 700;
                text-align: left;
            }
            .evaluation-row > div:nth-child(2) {
                color: #7c3aed;
            }
            .evaluation-row > div:nth-child(3) {
                color: #0f172a;
            }
            .evaluation-header-row {
                background: #f8fafc;
            }
            .evaluation-header-row > div {
                color: #64748b;
                font-size: 9px;
                font-weight: 800;
                letter-spacing: 0.03em;
                text-transform: uppercase;
            }
            .analytics-time-range {
                display: grid;
                grid-template-columns: repeat(4, 38px);
                gap: 3px;
                padding: 3px;
                border: 1px solid rgba(148, 163, 184, 0.28);
                border-radius: 9px;
                background: #f8fafc;
            }
            .analytics-time-range label {
                display: flex !important;
                align-items: center;
                justify-content: center;
                height: 28px;
                margin: 0 !important;
                border: 1px solid transparent;
                border-radius: 6px;
                color: #64748b;
                font-size: 10px;
                font-weight: 700;
                cursor: pointer;
                box-sizing: border-box;
            }
            .analytics-time-range input {
                display: none !important;
            }
            .analytics-time-range label:has(input:checked) {
                border-color: rgba(79, 70, 229, 0.32);
                background: #ffffff;
                color: #4f46e5;
                box-shadow: 0 2px 6px rgba(15, 23, 42, 0.07);
            }
            .summary-dashboard {
                display: flex;
                flex-direction: column;
                gap: 14px;
                min-width: 0;
            }
            .summary-hero-grid {
                display: grid;
                grid-template-columns: repeat(2, minmax(0, 1fr));
                gap: 10px;
            }
            .summary-hero {
                min-width: 0;
                padding: 11px 13px;
                border: 1px solid #e2e8f0;
                border-radius: 12px;
                background: #f8fafc;
                box-sizing: border-box;
            }
            .summary-hero.summary-tone-positive {
                border-color: rgba(22, 163, 74, 0.24);
                background: rgba(240, 253, 244, 0.72);
            }
            .summary-hero.summary-tone-negative {
                border-color: rgba(220, 38, 38, 0.22);
                background: rgba(254, 242, 242, 0.72);
            }
            .summary-hero.summary-tone-accent {
                border-color: rgba(79, 70, 229, 0.22);
                background: rgba(238, 242, 255, 0.72);
            }
            .summary-hero-label {
                color: #64748b;
                font-size: 9px;
                font-weight: 800;
                letter-spacing: 0.09em;
                text-transform: uppercase;
            }
            .summary-hero-value {
                margin-top: 3px;
                color: #0f172a;
                font-size: 23px;
                font-weight: 700;
                line-height: 1.1;
            }
            .summary-hero.summary-tone-positive .summary-hero-value,
            .summary-tone-positive {
                color: #15803d;
            }
            .summary-hero.summary-tone-negative .summary-hero-value,
            .summary-tone-negative {
                color: #b91c1c;
            }
            .summary-hero.summary-tone-accent .summary-hero-value,
            .summary-tone-accent {
                color: #4f46e5;
            }
            .summary-hero-subtitle {
                margin-top: 4px;
                overflow: hidden;
                color: #94a3b8;
                font-size: 9px;
                text-overflow: ellipsis;
                white-space: nowrap;
            }
            .summary-section-grid {
                display: grid;
                grid-template-columns: minmax(0, 1.25fr) minmax(0, 0.75fr);
                min-width: 0;
                border-top: 1px solid #e2e8f0;
                padding-top: 10px;
            }
            .summary-section {
                min-width: 0;
                padding: 0 12px;
            }
            .summary-section:first-child {
                padding-left: 0;
                border-right: 1px solid #e2e8f0;
            }
            .summary-section:last-child {
                padding-right: 0;
            }
            .summary-section-title {
                margin-bottom: 4px;
                color: #334155;
                font-size: 10px;
                font-weight: 800;
                letter-spacing: 0.06em;
                text-transform: uppercase;
            }
            .summary-stat {
                display: flex;
                align-items: center;
                justify-content: space-between;
                gap: 8px;
                min-height: 25px;
                border-bottom: 1px solid rgba(226, 232, 240, 0.72);
            }
            .summary-stat:last-child {
                border-bottom: 0;
            }
            .summary-stat-label {
                overflow: hidden;
                color: #64748b;
                font-size: 10px;
                text-overflow: ellipsis;
                white-space: nowrap;
            }
            .summary-stat-value {
                flex: 0 0 auto;
                color: #0f172a;
                font-size: 11px;
                font-weight: 700;
            }
            .primary-action-button,
            .secondary-action-button,
            .danger-action-button {
                min-height: 40px;
                padding: 0 16px;
                border-radius: 10px;
                font-family: Arial, sans-serif;
                font-size: 12px;
                font-weight: 700;
                cursor: pointer;
            }
            .trade-header-actions {
                display: flex;
                align-items: center;
                gap: 7px;
            }
            .primary-action-button {
                border: 1px solid #4f46e5;
                background: #4f46e5;
                color: #ffffff;
                box-shadow: 0 7px 16px rgba(79, 70, 229, 0.20);
            }
            .secondary-action-button {
                border: 1px solid #cbd5e1;
                background: #ffffff;
                color: #475569;
            }
            .danger-action-button {
                border: 1px solid rgba(220, 38, 38, 0.34);
                background: #ffffff;
                color: #b91c1c;
            }
            .danger-action-button:disabled {
                border-color: #e2e8f0;
                color: #94a3b8;
                cursor: not-allowed;
                opacity: 0.72;
            }
            .trade-modal-overlay {
                position: fixed;
                inset: 0;
                z-index: 1000;
                align-items: center;
                justify-content: center;
                padding: 20px;
                background: rgba(15, 23, 42, 0.48);
                backdrop-filter: blur(4px);
                box-sizing: border-box;
            }
            .trade-modal {
                width: min(760px, 100%);
                max-height: calc(100vh - 40px);
                overflow-y: auto;
                padding: 20px;
                border: 1px solid rgba(148, 163, 184, 0.35);
                border-radius: 18px;
                background: #ffffff;
                box-shadow: 0 26px 70px rgba(15, 23, 42, 0.28);
                box-sizing: border-box;
            }
            .trade-modal-header,
            .trade-modal-actions {
                display: flex;
                align-items: center;
                justify-content: space-between;
                gap: 10px;
            }
            .trade-modal-header {
                margin-bottom: 16px;
            }
            .trade-modal-title {
                color: #0f172a;
                font-size: 20px;
                font-weight: 700;
            }
            .modal-close-button {
                width: 34px;
                height: 34px;
                padding: 0;
                border: 1px solid #e2e8f0;
                border-radius: 9px;
                background: #f8fafc;
                color: #475569;
                font-size: 22px;
                line-height: 1;
                cursor: pointer;
            }
            .trade-modal-actions {
                justify-content: flex-end;
                margin-top: 14px;
            }
            .trade-modal .DateInput,
            .trade-modal .DateInput_input,
            .trade-modal .SingleDatePicker,
            .trade-modal .SingleDatePickerInput {
                width: 100%;
                box-sizing: border-box;
            }
            .form-field {
                display: flex;
                flex-direction: column;
                gap: 6px;
                padding: 12px;
                border-radius: 12px;
                background: rgba(255,255,255,0.8);
                border: 1px solid rgba(148, 163, 184, 0.25);
            }
            .form-field label,
            .control-row label {
                font-size: 12px;
                font-weight: 600;
                color: #475569;
            }
            input,
            .Select-control,
            .Select-menu-outer {
                min-height: 38px;
                box-sizing: border-box;
                border: 1px solid #cbd5e1;
                border-radius: 10px;
                background: white;
                padding: 0 10px;
                font-size: 14px;
            }
            .Select-control {
                border-radius: 10px;
                min-height: 38px;
                padding: 0;
            }
            input:focus {
                outline: none;
                border-color: #6366f1;
                box-shadow: 0 0 0 3px rgba(99, 102, 241, 0.12);
            }
            .time-range-card {
                display: flex;
                position: absolute;
                top: 12px;
                right: 12px;
                left: auto;
                transform: none;
                z-index: 30;
                align-items: center;
                justify-content: flex-end;
                gap: 0;
                width: auto;
                max-width: calc(100% - 24px);
                margin: 0;
                padding: 4px;
                border: 1px solid rgba(148, 163, 184, 0.28);
                border-radius: 11px;
                background: rgba(255, 255, 255, 0.90);
                box-shadow: 0 4px 14px rgba(15, 23, 42, 0.05);
                box-sizing: border-box;
            }
            .chart-panel .modebar {
                top: 48px !important;
                right: 8px !important;
            }
            .time-range-card .indicator-title {
                color: #94a3b8;
                font-size: 9px;
                letter-spacing: 0.16em;
                white-space: nowrap;
            }
            .time-range-buttons {
                display: grid;
                grid-template-columns: repeat(5, minmax(42px, 1fr));
                grid-auto-rows: 34px;
                gap: 4px;
                min-width: 0;
                margin: 0;
                padding: 0;
            }
            .time-range-buttons label {
                font-family: Arial, sans-serif !important;
                position: relative;
                display: flex !important;
                align-items: center;
                justify-content: center;
                height: 34px;
                margin: 0 !important;
                padding: 0 10px;
                border: 1px solid rgba(148, 163, 184, 0.30);
                border-radius: 7px;
                background: rgba(248, 250, 252, 0.78);
                color: #64748b;
                font-size: 11px;
                font-weight: 800;
                line-height: 1;
                cursor: pointer;
                box-sizing: border-box;
                transition: color 0.15s ease, border-color 0.15s ease, background 0.15s ease;
            }
            .time-range-buttons label:hover {
                color: #4f46e5;
                border-color: rgba(99, 102, 241, 0.42);
                background: rgba(238, 242, 255, 0.78);
            }
            .time-range-buttons input {
                display: none !important;
            }
            .time-range-buttons label:has(input:checked) {
                border-color: rgba(79, 70, 229, 0.52);
                background: rgba(238, 242, 255, 0.94);
                color: #4f46e5;
                box-shadow: inset 0 0 0 1px rgba(79, 70, 229, 0.08);
            }
            .regression-button,
            .bollinger-button,
            .kalman-button {
                width: 100%;
                margin-bottom: 14px;
            }
            .compact-label {
                display: block;
                margin: 0 0 6px;
                color: #475569;
                font-size: 11px;
                font-weight: 700;
            }
            .compact-input {
                display: flex;
                align-items: center;
                overflow: hidden;
                border: 1px solid #cbd5e1;
                border-radius: 11px;
                background: white;
            }
            .compact-input input {
                flex: 1 1 auto;
                min-width: 0;
                border: 0;
                box-shadow: none;
            }
            .input-suffix {
                padding: 0 12px;
                color: #64748b;
                font-size: 11px;
                font-weight: 700;
            }
            .regression-settings,
            .bollinger-settings,
            .kalman-settings {
                display: grid;
                grid-template-columns: repeat(2, minmax(0, 1fr));
                gap: 9px;
            }
            .regression-settings {
                grid-template-columns: repeat(3, minmax(0, 1fr));
            }
            .parameter-label-row {
                display: flex;
                align-items: center;
                gap: 6px;
                margin-bottom: 6px;
            }
            .parameter-label-row .compact-label {
                margin: 0;
            }
            .parameter-info {
                position: relative;
                display: inline-flex;
                align-items: center;
                justify-content: center;
                width: 20px;
                height: 20px;
                border: 1px solid #cbd5e1;
                border-radius: 50%;
                background: #f8fafc;
                color: #475569;
                cursor: help;
                outline: none;
            }
            .parameter-info-glyph {
                font-size: 11px;
                font-weight: 700;
                line-height: 1;
            }
            .parameter-tooltip {
                position: absolute;
                right: 0;
                bottom: calc(100% + 8px);
                z-index: 100;
                width: 210px;
                padding: 9px 11px;
                border-radius: 9px;
                background: #0f172a;
                color: #ffffff;
                font-size: 10px;
                font-weight: 500;
                line-height: 1.35;
                text-align: left;
                box-shadow: 0 8px 22px rgba(15, 23, 42, 0.22);
                opacity: 0;
                visibility: hidden;
                transform: translateY(4px);
                transition: opacity 0.15s ease, transform 0.15s ease, visibility 0.15s ease;
                pointer-events: none;
            }
            .parameter-info:hover .parameter-tooltip,
            .parameter-info:focus .parameter-tooltip {
                opacity: 1;
                visibility: visible;
                transform: translateY(0);
            }
            .setting-field {
                min-width: 0;
            }
            .setting-field input {
                width: 100%;
                min-width: 0;
            }
            .bollinger-button label {
                font-family: Arial, sans-serif !important;
                position: relative;
                display: flex !important;
                align-items: center;
                justify-content: center;
                min-height: 42px;
                padding: 8px 14px;
                border: 1px solid #2563eb;
                border-radius: 11px;
                background: white;
                color: #2563eb;
                font-size: 12px;
                font-weight: 700;
                cursor: pointer;
                user-select: none;
                transition: transform 0.15s ease, box-shadow 0.15s ease, background 0.15s ease;
                box-sizing: border-box;
            }
            .bollinger-button label:hover {
                transform: translateY(-1px);
                box-shadow: 0 7px 16px rgba(37, 99, 235, 0.16);
            }
            .bollinger-button input {
                display: none !important;
            }
            .bollinger-button label:has(input:checked) {
                color: white;
                background: linear-gradient(135deg, #2563eb, #3b82f6);
                box-shadow: 0 8px 18px rgba(37, 99, 235, 0.24);
            }
            .regression-button label {
                font-family: Arial, sans-serif !important;
                position: relative;
                display: flex !important;
                align-items: center;
                justify-content: center;
                min-height: 42px;
                padding: 8px 14px;
                border: 1px solid #7c3aed;
                border-radius: 11px;
                background: white;
                color: #7c3aed;
                font-size: 12px;
                font-weight: 700;
                cursor: pointer;
                user-select: none;
                transition: transform 0.15s ease, box-shadow 0.15s ease, background 0.15s ease;
                box-sizing: border-box;
            }
            .regression-button label:hover {
                transform: translateY(-1px);
                box-shadow: 0 7px 16px rgba(124, 58, 237, 0.16);
            }
            .regression-button input {
                display: none !important;
            }
            .regression-button label:has(input:checked) {
                color: white;
                background: linear-gradient(135deg, #7c3aed, #a855f7);
                box-shadow: 0 8px 18px rgba(124, 58, 237, 0.24);
            }
            .kalman-button label {
                font-family: Arial, sans-serif !important;
                position: relative;
                display: flex !important;
                align-items: center;
                justify-content: center;
                min-height: 42px;
                padding: 8px 14px;
                border: 1px solid #db2777;
                border-radius: 11px;
                background: white;
                color: #db2777;
                font-size: 12px;
                font-weight: 700;
                cursor: pointer;
                user-select: none;
                transition: transform 0.15s ease, box-shadow 0.15s ease, background 0.15s ease;
                box-sizing: border-box;
            }
            .kalman-button label:hover {
                transform: translateY(-1px);
                box-shadow: 0 7px 16px rgba(219, 39, 119, 0.16);
            }
            .kalman-button input {
                display: none !important;
            }
            .kalman-button label:has(input:checked) {
                color: white;
                background: linear-gradient(135deg, #be185d, #ec4899);
                box-shadow: 0 8px 18px rgba(219, 39, 119, 0.24);
            }
            button {
                padding: 10px 16px;
                cursor: pointer;
                margin-top: 6px;
                border: none;
                border-radius: 10px;
                background: linear-gradient(135deg, #4f46e5 0%, #7c3aed 100%);
                color: white;
                font-weight: 600;
                box-shadow: 0 8px 18px rgba(79, 70, 229, 0.2);
            }
            button:hover {
                filter: brightness(1.04);
            }
            .message {
                min-height: 22px;
                margin-top: 10px;
                color: #334155;
                font-size: 13px;
                font-weight: 500;
            }
            .metric-grid {
                display: flex;
                flex-wrap: wrap;
                gap: 12px;
                margin-bottom: 16px;
            }
            .chart-panel,
            .chart-panel > #main-chart,
            .chart-panel .js-plotly-plot,
            .chart-panel .plot-container {
                width: 100% !important;
                max-width: 100% !important;
                min-width: 0;
            }
            @media (max-width: 1250px) {
                .chart-workspace {
                    grid-template-columns: minmax(0, 1fr);
                    grid-template-areas:
                        "indicators"
                        "chart";
                }
                .indicator-sidebar {
                    grid-template-columns: repeat(3, minmax(0, 1fr));
                    width: 100%;
                }
            }
            @media (max-width: 900px) {
                .analytics-grid {
                    grid-template-columns: 1fr;
                }
                .analytics-chart-panel,
                .analytics-side-column {
                    height: 680px;
                }
                .time-range-card {
                    top: 10px;
                }
                .time-range-buttons {
                    grid-template-columns: repeat(5, minmax(40px, 1fr));
                }
            }
            @media (max-width: 680px) {
                .analytics-chart-header {
                    align-items: flex-start;
                    flex-direction: column;
                }
                .trade-form,
                .analytics-metrics {
                    grid-template-columns: 1fr;
                }
                .trade-modal-overlay {
                    padding: 10px;
                }
                .time-range-card .indicator-title {
                    display: none;
                }
                .indicator-sidebar {
                    grid-template-columns: 1fr;
                }
            }
        </style>
    </head>
    <body>
        {%app_entry%}
        <footer>
            {%config%}
            {%scripts%}
            {%renderer%}
        </footer>
    </body>
</html>
"""


app.clientside_callback(
    """
    function(figure) {
        window.setTimeout(function() {
            const root = document.getElementById("main-chart");
            const graph = root && root.querySelector(".js-plotly-plot");
            if (!graph || !graph._fullLayout || typeof graph.on !== "function") {
                return;
            }

            if (graph.__msciHoverHandler && typeof graph.removeListener === "function") {
                graph.removeListener("plotly_hover", graph.__msciHoverHandler);
                graph.removeListener("plotly_unhover", graph.__msciUnhoverHandler);
            }

            let line = graph.querySelector(".cross-panel-hover-line");
            if (!line) {
                line = document.createElement("div");
                line.className = "cross-panel-hover-line";
                Object.assign(line.style, {
                    position: "absolute",
                    display: "none",
                    width: "0",
                    borderLeft: "1px dashed rgba(37, 99, 235, 0.58)",
                    pointerEvents: "none",
                    zIndex: "20"
                });
                graph.appendChild(line);
            }

            graph.__msciHoverHandler = function(eventData) {
                const point = eventData && eventData.points && eventData.points[0];
                const size = graph._fullLayout && graph._fullLayout._size;
                if (!point || !point.xaxis || !size) {
                    line.style.display = "none";
                    return;
                }
                const xPixel = point.xaxis.d2p(point.x) + point.xaxis._offset;
                line.style.left = xPixel + "px";
                line.style.top = size.t + "px";
                line.style.height = size.h + "px";
                line.style.display = "block";
            };
            graph.__msciUnhoverHandler = function() {
                line.style.display = "none";
            };
            graph.on("plotly_hover", graph.__msciHoverHandler);
            graph.on("plotly_unhover", graph.__msciUnhoverHandler);
        }, 0);
        return Date.now();
    }
    """,
    Output("hover-sync-token", "data"),
    Input("main-chart", "figure"),
)


@app.callback(
    Output("instrument-meta", "children"),
    Input("instrument-selector", "value"),
)
def update_instrument_meta(instrument_key):
    config = instrument_config(instrument_key)
    return (
        f"{config['name']} · ISIN {config['isin']} · Yahoo {config['ticker']} · "
        "Trading Tools: Tagesdaten (MAX, Adj Close) · Trading Analytics: 5-Minuten-Daten"
    )


@app.callback(
    Output("market-reload-message", "children"),
    Output("market-data-refresh-token", "data"),
    Input("reload-market-data-confirm", "submit_n_clicks"),
    State("market-data-refresh-token", "data"),
    State("instrument-selector", "value"),
    prevent_initial_call=True,
)
def reload_market_data(_submit_clicks, refresh_token, instrument_key):
    selected_key = instrument_key or DEFAULT_INSTRUMENT_KEY
    results = []
    errors = []
    refresh_jobs = [
        (
            "Daily",
            instrument_sheet_name(selected_key, "daily"),
            DAILY_PERIOD,
            DAILY_INTERVAL,
            instrument_market_cache(selected_key, "daily"),
        ),
        (
            "5 min",
            instrument_sheet_name(selected_key, "intraday"),
            INTRADAY_PERIOD,
            INTRADAY_INTERVAL,
            instrument_market_cache(selected_key, "intraday"),
        ),
    ]
    for label, sheet_name, period, interval, cache in refresh_jobs:
        try:
            added = refresh_market_history(
                instrument_key=selected_key,
                sheet_name=sheet_name,
                period=period,
                interval=interval,
                cache=cache,
                force=True,
            )
            results.append(f"{label}: {added} neue Datenpunkte")
        except Exception as exc:
            errors.append(f"{label}: {exc}")

    try:
        _timestamp, last_price = load_last_price(selected_key, force=True)
        results.append(f"Last Price: {last_price:.4f}")
    except Exception as exc:
        errors.append(f"Last Price: {exc}")

    message = " · ".join(results)
    if errors:
        message = f"{message} · Fehler: {'; '.join(errors)}" if message else f"Fehler: {'; '.join(errors)}"
    return message, int(refresh_token or 0) + 1


@app.callback(
    Output("main-chart", "figure"),
    Input("regression-toggle", "value"),
    Input("reg-6m-days", "value"),
    Input("reg-1y-days", "value"),
    Input("reg-2y-days", "value"),
    Input("bollinger-toggle", "value"),
    Input("bollinger-window", "value"),
    Input("bollinger-std", "value"),
    Input("kalman-toggle", "value"),
    Input("kalman-process-variance", "value"),
    Input("kalman-measurement-variance", "value"),
    Input("time-range-selector", "value"),
    Input("main-chart", "relayoutData"),
    Input("instrument-selector", "value"),
    Input("refresh-token", "data"),
    Input("market-data-refresh-token", "data"),
)
def refresh_dashboard(
    regression_toggle,
    reg_6m_days,
    reg_1y_days,
    reg_2y_days,
    bollinger_toggle,
    bollinger_window,
    bollinger_std,
    kalman_toggle,
    kalman_process_variance,
    kalman_measurement_variance,
    time_range_selection,
    relayout_data,
    instrument_key,
    _refresh_token,
    _market_data_refresh_token,
):
    # Trading Tools receives the latest quote. Trading Analytics calls
    # load_market_data() separately and remains based on 5-minute bars.
    selected_key = instrument_key or DEFAULT_INSTRUMENT_KEY
    market_df = load_trading_tools_data(selected_key)
    trades = (
        load_trades()
        if selected_key == DEFAULT_INSTRUMENT_KEY
        else pd.DataFrame(columns=TRADE_COLUMNS)
    )

    settings = {
        "regression_windows": list(
            dict.fromkeys(
                [
                    _positive_int_input(reg_6m_days, 182, minimum=2),
                    _positive_int_input(reg_1y_days, 365, minimum=2),
                    _positive_int_input(reg_2y_days, 1825, minimum=2),
                ]
            )
        ),
        "show_regression": "show" in (regression_toggle or []),
        "bollinger_window": _positive_int_input(
            bollinger_window, DEFAULT_SETTINGS["bollinger_window"], minimum=2
        ),
        "bollinger_std": _positive_float_input(
            bollinger_std, DEFAULT_SETTINGS["bollinger_std"]
        ),
        "show_bollinger": "show" in (bollinger_toggle or []),
        "kalman_process_variance": _positive_float_input(
            kalman_process_variance, DEFAULT_KALMAN_PROCESS_VARIANCE
        ),
        "kalman_measurement_variance": _positive_float_input(
            kalman_measurement_variance, DEFAULT_KALMAN_MEASUREMENT_VARIANCE
        ),
        "show_kalman": "show" in (kalman_toggle or []),
    }

    control_range = visible_range_from_control(
        market_df,
        time_range_selection,
    )
    try:
        triggered_id = ctx.triggered_id
    except Exception:
        triggered_id = None
    if triggered_id == "main-chart":
        relayout_range = visible_range_from_relayout(relayout_data)
        visible_range = (
            relayout_range
            if relayout_range is not None
            else control_range if time_range_selection != "max" else None
        )
    elif triggered_id == "time-range-selector":
        visible_range = control_range
    elif time_range_selection and time_range_selection != "max":
        visible_range = control_range
    else:
        visible_range = visible_range_from_relayout(relayout_data)

    figure = build_figure(
        market_df,
        settings,
        trades,
        visible_range=visible_range,
        instrument_name=instrument_config(selected_key)["name"],
    )

    return figure


def indexed_investment_series(
    market_df: pd.DataFrame, trades: pd.DataFrame
) -> pd.DataFrame:
    chart_df = market_df[["Date", "Price"]].dropna().sort_values("Date")
    if trades.empty or trades["Entry_Date"].dropna().empty:
        return pd.DataFrame(
            columns=["Date", "Strategy_Index", "Buy_Hold_Index", "Invested"]
        )

    first_entry = pd.Timestamp(trades["Entry_Date"].dropna().min())
    comparison = chart_df[chart_df["Date"] >= first_entry].copy()
    if comparison.empty:
        return pd.DataFrame(
            columns=["Date", "Strategy_Index", "Buy_Hold_Index", "Invested"]
        )

    dates = comparison["Date"]
    invested = np.zeros(len(comparison), dtype=bool)
    for trade in trades.dropna(subset=["Entry_Date"]).itertuples(index=False):
        entry_date = pd.Timestamp(trade.Entry_Date)
        exit_date = (
            pd.Timestamp(trade.Exit_Date) if pd.notna(trade.Exit_Date) else None
        )
        active = (dates >= entry_date).to_numpy(copy=True)
        if exit_date is not None:
            active &= (dates < exit_date).to_numpy()
        invested |= active

    market_returns = comparison["Price"].pct_change().fillna(0.0)
    exposure_for_return = pd.Series(invested, index=comparison.index).shift(
        1, fill_value=False
    )
    strategy_returns = market_returns.where(exposure_for_return, 0.0)
    comparison["Buy_Hold_Index"] = 100.0 * (1.0 + market_returns).cumprod()
    comparison["Strategy_Index"] = 100.0 * (1.0 + strategy_returns).cumprod()
    comparison["Invested"] = invested
    return comparison[
        ["Date", "Strategy_Index", "Buy_Hold_Index", "Invested"]
    ]


def build_trade_analytics_figure(
    market_df: pd.DataFrame, trades: pd.DataFrame, time_range: str = "max"
) -> go.Figure:
    chart_df = market_df[["Date", "Price"]].dropna().sort_values("Date")
    data_end = pd.Timestamp(chart_df["Date"].max())
    market_start = pd.Timestamp(chart_df["Date"].min())
    comparison = indexed_investment_series(market_df, trades)
    if comparison.empty and not chart_df.empty:
        comparison = chart_df.copy()
        returns = comparison["Price"].pct_change().fillna(0.0)
        comparison["Buy_Hold_Index"] = 100.0 * (1.0 + returns).cumprod()
        comparison["Strategy_Index"] = np.nan
        comparison["Invested"] = False
    comparison_start = (
        pd.Timestamp(comparison["Date"].min()) if not comparison.empty else market_start
    )
    offsets = {
        "1y": pd.DateOffset(years=1),
        "2y": pd.DateOffset(years=2),
        "5y": pd.DateOffset(years=5),
    }
    data_start = (
        max(comparison_start, data_end - offsets[time_range])
        if time_range in offsets
        else comparison_start
    )
    visible_comparison = comparison[
        (comparison["Date"] >= data_start) & (comparison["Date"] <= data_end)
    ]

    figure = go.Figure()
    if not visible_comparison.empty:
        figure.add_trace(
            go.Scatter(
                x=visible_comparison["Date"],
                y=visible_comparison["Buy_Hold_Index"],
                mode="lines",
                name="Buy & Hold",
                line={"color": "#0f172a", "width": 2.3},
                hovertemplate="Buy & Hold: %{y:.2f}<extra></extra>",
            )
        )
        if visible_comparison["Strategy_Index"].notna().any():
            figure.add_trace(
                go.Scatter(
                    x=visible_comparison["Date"],
                    y=visible_comparison["Strategy_Index"],
                    mode="lines",
                    name="Meine Strategie",
                    line={"color": "#7c3aed", "width": 2.3},
                    hovertemplate="Meine Strategie: %{y:.2f}<extra></extra>",
                )
            )

    if not visible_comparison.empty:
        interval_start = None
        previous_date = None
        for point in visible_comparison.itertuples(index=False):
            current_date = pd.Timestamp(point.Date)
            if point.Invested and interval_start is None:
                interval_start = current_date
            elif not point.Invested and interval_start is not None:
                figure.add_vrect(
                    x0=interval_start,
                    x1=current_date,
                    fillcolor="rgba(22, 163, 74, 0.055)",
                    line_width=0,
                    layer="below",
                )
                interval_start = None
            previous_date = current_date
        if interval_start is not None and previous_date is not None:
            figure.add_vrect(
                x0=interval_start,
                x1=previous_date,
                fillcolor="rgba(22, 163, 74, 0.055)",
                line_width=0,
                layer="below",
            )

    if not trades.empty and not comparison.empty:
        def index_at(dates: pd.Series) -> list[float]:
            return [
                float(
                    comparison.iloc[
                        (comparison["Date"] - pd.Timestamp(date_value)).abs().argmin()
                    ]["Buy_Hold_Index"]
                )
                for date_value in dates
            ]

        entry_points = trades.dropna(subset=["Entry_Date"]).copy()
        entry_points = entry_points[
            (entry_points["Entry_Date"] >= data_start)
            & (entry_points["Entry_Date"] <= data_end)
        ]
        if not entry_points.empty:
            figure.add_trace(
                go.Scatter(
                    x=entry_points["Entry_Date"],
                    y=index_at(entry_points["Entry_Date"]),
                    mode="markers",
                    name="Einstieg",
                    showlegend=False,
                    marker={"symbol": "triangle-up", "size": 12, "color": "#16a34a", "line": {"width": 1.3, "color": "#ffffff"}},
                    customdata=entry_points["Trade_ID"].astype(str),
                    hovertemplate="Einstieg · Trade %{customdata}<br>%{x|%d.%m.%Y}<extra></extra>",
                )
            )

        exit_points = trades.dropna(subset=["Exit_Date"]).copy()
        exit_points = exit_points[
            (exit_points["Exit_Date"] >= data_start)
            & (exit_points["Exit_Date"] <= data_end)
        ]
        if not exit_points.empty:
            figure.add_trace(
                go.Scatter(
                    x=exit_points["Exit_Date"],
                    y=index_at(exit_points["Exit_Date"]),
                    mode="markers",
                    name="Ausstieg",
                    showlegend=False,
                    marker={"symbol": "triangle-down", "size": 12, "color": "#dc2626", "line": {"width": 1.3, "color": "#ffffff"}},
                    customdata=exit_points["Trade_ID"].astype(str),
                    hovertemplate="Ausstieg · Trade %{customdata}<br>%{x|%d.%m.%Y}<extra></extra>",
                )
            )

    figure.update_layout(
        template="plotly_white",
        autosize=True,
        uirevision=f"trading-analytics-{time_range}",
        hovermode="x unified",
        showlegend=True,
        paper_bgcolor="#ffffff",
        plot_bgcolor="#ffffff",
        font={"family": "Arial, sans-serif", "color": "#0f172a"},
        margin={"l": 52, "r": 18, "t": 50, "b": 42},
        legend={"orientation": "h", "x": 0, "y": 1.08, "xanchor": "left", "yanchor": "bottom", "font": {"size": 10}, "bgcolor": "rgba(255,255,255,0)"},
    )
    axis_style = {
        "showgrid": False,
        "showline": False,
        "ticks": "",
        "tickfont": {"size": 10, "color": "#64748b"},
        "automargin": True,
    }
    figure.update_xaxes(
        range=[data_start, data_end],
        hoverformat="%d.%m.%Y",
        rangeslider_visible=False,
        rangebreaks=trading_day_rangebreaks(chart_df),
        **axis_style,
    )
    indexed_values = pd.concat(
        [
            visible_comparison["Buy_Hold_Index"],
            visible_comparison["Strategy_Index"],
        ],
        ignore_index=True,
    )
    figure.update_yaxes(
        range=padded_axis_range(indexed_values),
        zeroline=False,
        title_text=None,
        tickformat=".0f",
        **axis_style,
    )
    return figure


def indexed_performance_statistics(
    dates: pd.Series, index_values: pd.Series
) -> dict[str, float]:
    values = pd.to_numeric(index_values, errors="coerce").dropna()
    if len(values) < 2:
        return {key: np.nan for key in ("cagr", "volatility", "sharpe", "max_drawdown")}
    aligned_dates = pd.to_datetime(dates.loc[values.index])
    calendar_days = max((aligned_dates.iloc[-1] - aligned_dates.iloc[0]).days, 1)
    returns = values.pct_change().dropna()
    volatility = float(returns.std(ddof=1) * np.sqrt(252)) if len(returns) > 1 else np.nan
    sharpe = (
        float(returns.mean() / returns.std(ddof=1) * np.sqrt(252))
        if len(returns) > 1 and returns.std(ddof=1) > 0
        else np.nan
    )
    running_max = values.cummax()
    max_drawdown = float((values / running_max - 1.0).min())
    cagr = float((values.iloc[-1] / values.iloc[0]) ** (365.25 / calendar_days) - 1.0)
    return {
        "cagr": cagr,
        "volatility": volatility,
        "sharpe": sharpe,
        "max_drawdown": max_drawdown,
    }


def evaluation_table(market_df: pd.DataFrame, trades: pd.DataFrame) -> html.Div:
    comparison = indexed_investment_series(market_df, trades)
    if comparison.empty:
        rows = [
            ("Rendite p.a.", "–", "–", "–"),
            ("Volatilität p.a.", "–", "–", "–"),
            ("Sharpe Ratio", "–", "–", "–"),
            ("Max. Drawdown", "–", "–", "–"),
            ("Information Ratio", "–", "–", "–"),
        ]
    else:
        strategy_stats = indexed_performance_statistics(
            comparison["Date"], comparison["Strategy_Index"]
        )
        benchmark_stats = indexed_performance_statistics(
            comparison["Date"], comparison["Buy_Hold_Index"]
        )
        strategy_returns = comparison["Strategy_Index"].pct_change().dropna()
        benchmark_returns = comparison["Buy_Hold_Index"].pct_change().dropna()
        active_returns = strategy_returns - benchmark_returns
        information_ratio = (
            float(active_returns.mean() / active_returns.std(ddof=1) * np.sqrt(252))
            if len(active_returns) > 1 and active_returns.std(ddof=1) > 0
            else np.nan
        )

        def percentage(value: float) -> str:
            return "–" if pd.isna(value) else f"{value * 100:.2f} %"

        def number(value: float) -> str:
            return "–" if pd.isna(value) else f"{value:.2f}"

        rows = [
            (
                "Rendite p.a.",
                percentage(strategy_stats["cagr"]),
                percentage(benchmark_stats["cagr"]),
                percentage(strategy_stats["cagr"] - benchmark_stats["cagr"]),
            ),
            (
                "Volatilität p.a.",
                percentage(strategy_stats["volatility"]),
                percentage(benchmark_stats["volatility"]),
                percentage(strategy_stats["volatility"] - benchmark_stats["volatility"]),
            ),
            (
                "Sharpe Ratio",
                number(strategy_stats["sharpe"]),
                number(benchmark_stats["sharpe"]),
                number(strategy_stats["sharpe"] - benchmark_stats["sharpe"]),
            ),
            (
                "Max. Drawdown",
                percentage(strategy_stats["max_drawdown"]),
                percentage(benchmark_stats["max_drawdown"]),
                percentage(strategy_stats["max_drawdown"] - benchmark_stats["max_drawdown"]),
            ),
            (
                "Information Ratio",
                number(information_ratio),
                "–",
                "–",
            ),
        ]

    return html.Div(
        [
            html.Div(
                [
                    html.Div("Kennzahl"),
                    html.Div("Meine Strategie"),
                    html.Div("Buy & Hold"),
                    html.Div("Differenz"),
                ],
                className="evaluation-row evaluation-header-row",
            ),
            *[
                html.Div(
                    [html.Div(label), html.Div(strategy), html.Div(benchmark), html.Div(difference)],
                    className="evaluation-row",
                )
                for label, strategy, benchmark, difference in rows
            ],
        ],
        className="evaluation-table",
    )


def trade_dashboard_payload(
    trades: pd.DataFrame | None = None, market_df: pd.DataFrame | None = None
):
    trades = load_trades() if trades is None else trades
    market_df = load_market_data() if market_df is None else market_df
    table_df = trades.copy()
    if not table_df.empty:
        for col in ["Entry_Date", "Exit_Date"]:
            table_df[col] = table_df[col].dt.strftime("%d.%m.%Y")
        for col in ["Gross_Return_Pct", "Net_Return_Pct", "Net_PnL"]:
            table_df[col] = table_df[col].round(2)

    column_labels = {
        "Trade_ID": "Trade",
        "Entry_Date": "Einstieg",
        "Exit_Date": "Ausstieg",
        "Direction": "Richtung",
        "Quantity": "Anzahl",
        "Entry_Price": "Einstiegskurs",
        "Exit_Price": "Ausstiegskurs",
        "Fees": "Gebühren",
        "Status": "Status",
        "Holding_Days": "Haltedauer",
        "Gross_Return_Pct": "Brutto %",
        "Net_Return_Pct": "Netto %",
        "Net_PnL": "Netto P&L",
        "Notes": "Notiz",
    }
    columns = [
        {"name": column_labels.get(col, col.replace("_", " ")), "id": col}
        for col in TRADE_COLUMNS
        if col not in {
            "Direction",
            "Quantity",
            "Status",
            "Gross_Return_Pct",
            "Net_PnL",
            "Net_Return_Pct",
        }
        and (col in table_df.columns or trades.empty)
    ]

    metrics = evaluation_table(market_df, trades)

    records = table_df.to_dict("records")
    for record in records:
        record["id"] = record.get("Trade_ID")
    return records, columns, metrics


@app.callback(
    Output("trade-table", "data"),
    Output("trade-table", "columns"),
    Output("trade-metrics", "children"),
    Output("analytics-chart", "figure"),
    Output("trade-table", "active_cell"),
    Input("refresh-token", "data"),
    Input("analytics-time-range", "value"),
    Input("market-data-refresh-token", "data"),
    Input("instrument-selector", "value"),
)
def refresh_trade_analytics(
    _refresh_token, time_range, _market_data_refresh_token, instrument_key
):
    selected_key = instrument_key or DEFAULT_INSTRUMENT_KEY
    trades = (
        load_trades()
        if selected_key == DEFAULT_INSTRUMENT_KEY
        else pd.DataFrame(columns=TRADE_COLUMNS)
    )
    market_df = load_market_data(selected_key)
    table_data, columns, metrics = trade_dashboard_payload(trades, market_df)
    figure = build_trade_analytics_figure(market_df, trades, time_range or "max")
    return table_data, columns, metrics, figure, None


@app.callback(
    Output("delete-trade", "disabled"),
    Input("trade-table", "active_cell"),
)
def toggle_delete_trade_button(active_cell):
    return not bool(active_cell and active_cell.get("row_id") is not None)


@app.callback(
    Output("settings-message", "children"),
    Input("save-settings", "n_clicks"),
    State("regression-toggle", "value"),
    State("reg-6m-days", "value"),
    State("reg-1y-days", "value"),
    State("reg-2y-days", "value"),
    State("bollinger-toggle", "value"),
    State("bollinger-window", "value"),
    State("bollinger-std", "value"),
    State("kalman-toggle", "value"),
    State("kalman-process-variance", "value"),
    State("kalman-measurement-variance", "value"),
    prevent_initial_call=True,
)
def persist_settings(
    _n_clicks,
    regression_toggle,
    reg_6m_days,
    reg_1y_days,
    reg_2y_days,
    bollinger_toggle,
    bollinger_window,
    bollinger_std,
    kalman_toggle,
    kalman_process_variance,
    kalman_measurement_variance,
):
    try:
        settings = {
            "regression_6m_days": _positive_int_input(reg_6m_days, 182, minimum=2),
            "regression_1y_days": _positive_int_input(reg_1y_days, 365, minimum=2),
            "regression_2y_days": _positive_int_input(reg_2y_days, 1825, minimum=2),
            "regression_5y_days": _positive_int_input(reg_2y_days, 1825, minimum=2),
            "show_regression_6m": "show" in (regression_toggle or []),
            "show_regression_1y": "show" in (regression_toggle or []),
            "show_regression_2y": "show" in (regression_toggle or []),
            "show_regression_5y": False,
            "show_regression_max": False,
            "bollinger_window": _positive_int_input(bollinger_window, 20, minimum=2),
            "bollinger_std": _positive_float_input(bollinger_std, 2.0),
            "show_bollinger": "show" in (bollinger_toggle or []),
            "kalman_process_variance": _positive_float_input(
                kalman_process_variance, DEFAULT_KALMAN_PROCESS_VARIANCE
            ),
            "kalman_measurement_variance": _positive_float_input(
                kalman_measurement_variance, DEFAULT_KALMAN_MEASUREMENT_VARIANCE
            ),
            "show_kalman": "show" in (kalman_toggle or []),
        }

        if settings["bollinger_window"] < 2:
            raise ValueError("Das Bollinger-Fenster muss mindestens 2 betragen.")
        if settings["bollinger_std"] <= 0:
            raise ValueError("Der Bollinger-Multiplikator muss größer als 0 sein.")

        save_settings(settings)
        return "Parameter wurden als Standardwerte im Setup-Sheet gespeichert."
    except Exception as exc:
        return f"Fehler: {exc}"


@app.callback(
    Output("trade-modal-overlay", "style"),
    Input("open-trade-modal", "n_clicks"),
    Input("close-trade-modal", "n_clicks"),
    Input("cancel-trade-modal", "n_clicks"),
    Input("refresh-token", "data"),
    prevent_initial_call=True,
)
def toggle_trade_modal(_open_clicks, _close_clicks, _cancel_clicks, _refresh_token):
    if ctx.triggered_id == "open-trade-modal":
        return {"display": "flex"}
    return {"display": "none"}


@app.callback(
    Output("trade-message", "children"),
    Output("refresh-token", "data"),
    Input("save-trade", "n_clicks"),
    Input("delete-trade-confirm", "submit_n_clicks"),
    State("trade-table", "active_cell"),
    State("entry-date", "date"),
    State("exit-date", "date"),
    State("entry-price", "value"),
    State("exit-price", "value"),
    State("fees", "value"),
    State("notes", "value"),
    State("refresh-token", "data"),
    prevent_initial_call=True,
)
def persist_trade(
    _save_clicks,
    _delete_clicks,
    active_cell,
    entry_date_value,
    exit_date_value,
    entry_price_value,
    exit_price_value,
    fees,
    notes,
    refresh_token,
):
    try:
        if ctx.triggered_id == "delete-trade-confirm":
            trade_id = (active_cell or {}).get("row_id")
            if trade_id is None:
                raise ValueError("Bitte zuerst einen Trade in der Tabelle auswählen.")
            message = delete_trade(trade_id)
            return message, int(refresh_token or 0) + 1

        if not entry_date_value:
            raise ValueError("Bitte ein Entry-Datum angeben.")
        market_df = load_market_data()
        message = append_trade(
            market_df=market_df,
            entry_date_value=entry_date_value,
            exit_date_value=exit_date_value,
            direction="Long",
            quantity=1.0,
            entry_price_value=entry_price_value,
            exit_price_value=exit_price_value,
            fees=float(fees or 0.0),
            notes=notes or "",
        )
        return message, int(refresh_token or 0) + 1
    except PermissionError:
        return (
            "Die Excel-Datei ist vermutlich in Excel geöffnet. Bitte schließen und erneut speichern.",
            no_update,
        )
    except Exception as exc:
        return f"Fehler: {exc}", no_update


if __name__ == "__main__":
    hot_reload = os.getenv("MSCI_HOT_RELOAD", "1").strip().lower() not in {
        "0",
        "false",
        "nein",
        "no",
    }
    app.run(
        debug=hot_reload,
        use_reloader=hot_reload,
        dev_tools_hot_reload=hot_reload,
        dev_tools_hot_reload_interval=1.0,
        dev_tools_hot_reload_watch_interval=0.5,
    )
'''

requirements = """dash>=3.0
pandas>=2.0
numpy>=1.26
plotly>=5.20
openpyxl>=3.1
yfinance>=0.2.65
"""

if __name__ == "__main__":
    exec(app_code, globals())
